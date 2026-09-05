import os
import json
import requests
import xml.etree.ElementTree as ET
import urllib.parse
from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
import google.generativeai as genai
import random
# dotenv removed

import base64
import threading
import uuid
from rule_engine import evaluate_knowledge_base, get_all_variables
from job_store import JobStore
try:
    from kcsc_mcp import kcsc_engine
except Exception as e:
    print(f"Failed to import kcsc_engine: {e}")
    kcsc_engine = None
from doc_extract import extract_text_from_file

app = Flask(__name__)
CORS(app)

# 파일(SQLite) 기반 저장소. 예전엔 순수 dict라 프로세스 재시작/멀티 워커 사이에서
# 작업 상태가 유실됐는데(운영에서 "존재하지 않는 작업입니다" 404로 재현됨),
# JobStore는 dict와 동일한 인터페이스를 제공하면서 jobs.db 파일에 저장한다.
JOBS = JobStore()
JOBS.cleanup_stale()


# API keys from environment variables
GEMINI_KEY = os.environ.get('GEMINI_API_KEY', '')

_V = "RDNDMEEyNTktQjQ1QS0zQ0U2LTg0MUQtNjJFRkIxMDNEM0NC"
VWORLD_KEY = os.environ.get('VWORLD_API_KEY', '') or base64.b64decode(_V).decode('utf-8')

_L = "a3NoODQ4MQ==" # ksh8481
LAW_KEY = os.environ.get('LAW_API_KEY', '') or base64.b64decode(_L).decode('utf-8')


def upload_file_for_gemini(path, display_name=None, timeout_seconds=90, poll_interval=2):
    """
    파일을 Gemini에 업로드한다. 반환된 객체는 .name/.uri/.mime_type을 갖고 있어서,
    - generate_content()에 바로 넣으려면: gemini_file_part(gfile) 로 dict로 변환해서 사용
    - 나중에(다른 요청에서) 다시 불러오려면: genai.get_file(gfile.name) (구버전 SDK로도 조회 가능)

    ⚠️ 예전엔 여기서 구버전(EOL 예정) `google.generativeai`의 genai.upload_file()을 썼는데,
    이 프로젝트의 GEMINI_API_KEY 형식이 그 구버전 업로드 경로(레거시 Google API Discovery
    클라이언트)와 호환이 안 돼서 "API key not valid" 에러로 조용히 실패하고 있었다. 문제는
    generate_content() 자체는 다른(더 최신) 경로를 타서 멀쩡히 동작했다는 것 - 그래서 텍스트
    분석은 잘 되는데 "파일 첨부"만 매번 조용히 실패하는 상황이 한동안 발견되지 않았다.
    특히 스캔본(이미지 전용) PDF처럼 로컬 텍스트 추출(fitz/PyPDF2)도 안 되는 파일은 업로드
    실패 시 대체 수단이 전혀 없어서 "파일을 통째로 못 읽는다"는 증상으로 드러났다.

    신버전 `google-genai` SDK(google.genai.Client)로 업로드하면 이 API 키로도 정상 동작하고,
    그 결과(file_uri/mime_type)를 구버전 SDK의 generate_content()에 {"file_data": {...}}
    형태로 그대로 넣어도 잘 읽는다는 걸 실제로 확인했다. 그래서 모델 폴백 루프 등 나머지
    코드는 손대지 않고 "업로드" 부분만 신버전 SDK로 교체한다.
    """
    import time
    from google import genai as google_genai_client

    client = google_genai_client.Client(api_key=GEMINI_KEY)
    gfile = client.files.upload(file=path, config={'display_name': display_name} if display_name else None)

    elapsed = 0
    while gfile.state.name == "PROCESSING" and elapsed < timeout_seconds:
        time.sleep(poll_interval)
        elapsed += poll_interval
        gfile = client.files.get(name=gfile.name)

    if gfile.state.name == "FAILED":
        raise Exception(f"Gemini 파일 처리 실패(FAILED): {display_name or path}")
    if gfile.state.name == "PROCESSING":
        print(f"[upload_file_for_gemini] {display_name or path} 파일이 {timeout_seconds}초 후에도 여전히 처리 중입니다. 그대로 진행합니다.")

    return gfile


def gemini_file_part(gfile):
    """upload_file_for_gemini()가 반환한 파일 객체를 generate_content() payload용 조각으로 변환."""
    return {"file_data": {"mime_type": gfile.mime_type, "file_uri": gfile.uri}}


def generate_with_search_grounding(prompt, model_name):
    """
    구글 실시간 검색(Google Search Grounding)을 사용해 generate_content를 실행한다.

    ⚠️ 예전엔 구버전(EOL 예정) google.generativeai에서 tools='google_search_retrieval'로
    검색 도구를 붙였는데, 그 도구 이름이 서버에서 이미 폐기돼서 실제 호출 시 매번
    "google_search_retrieval is not supported. Please use google_search tool instead."
    라는 400 에러가 100% 발생하고 있었다. run_analysis의 try/except가 이 실패를 조용히
    삼키고 검색 없이 일반 모델로만 폴백해서, law_review 기능은 지금까지 한 번도 실제
    구글 검색 그라운딩을 쓴 적이 없었다 (프롬프트는 "구글 검색 도구를 활용하라"고
    지시하고 있었지만 실제로는 매번 무시됐음).

    구버전 SDK는 최신 tools='google_search' 문법 자체를 모른다(SDK가 그 이후에 추가된
    기능이라 클라이언트 코드에 정의돼 있지 않음). 신버전 `google-genai` SDK로는 실제로
    검색 그라운딩이 정상 동작하는 걸 확인해서, 파일 업로드 때와 같은 방식으로 이 호출만
    신버전 SDK로 옮긴다. 반환값은 신버전 SDK의 응답 객체인데 .text 프로퍼티가 동일하게
    있어서, 이 함수를 호출하는 쪽 코드(response.text 사용)는 그대로 쓸 수 있다.
    """
    from google import genai as google_genai_client
    from google.genai import types

    client = google_genai_client.Client(api_key=GEMINI_KEY)
    grounding_tool = types.Tool(google_search=types.GoogleSearch())
    config = types.GenerateContentConfig(tools=[grounding_tool])

    clean_model_name = model_name.replace('models/', '')
    return client.models.generate_content(model=clean_model_name, contents=prompt, config=config)


SIDO_DATA = [
    {"code": "11", "name": "서울특별시"}, {"code": "26", "name": "부산광역시"},
    {"code": "27", "name": "대구광역시"}, {"code": "28", "name": "인천광역시"},
    {"code": "29", "name": "광주광역시"}, {"code": "30", "name": "대전광역시"},
    {"code": "31", "name": "울산광역시"}, {"code": "36", "name": "세종특별자치시"},
    {"code": "41", "name": "경기도"}, {"code": "43", "name": "충청북도"},
    {"code": "44", "name": "충청남도"}, {"code": "45", "name": "전북특별자치도"},
    {"code": "46", "name": "전라남도"}, {"code": "47", "name": "경상북도"},
    {"code": "48", "name": "경상남도"}, {"code": "50", "name": "제주특별자치도"},
    {"code": "51", "name": "강원특별자치도"}
]

@app.route('/api/regions/sido', methods=['GET'])
def get_sido():
    return jsonify({"success": True, "data": SIDO_DATA})

@app.route('/api/regions/<layer>', methods=['GET'])
def get_regions(layer):
    vworld_key = request.args.get('vworldKey')
    parent_code = request.args.get('parentCode')
    
    layer_map = {
        'sigungu': ('LT_C_ADSIGG_INFO', 'sig_cd', 'sig_kor_nm'),
        'emd': ('LT_C_ADEMD_INFO', 'emd_cd', 'emd_kor_nm'),
        'ri': ('LT_C_ADRI_INFO', 'li_cd', 'li_kor_nm')
    }
    
    if layer not in layer_map:
        return jsonify({"success": False, "message": "Invalid layer"})
        
    v_layer, code_field, name_field = layer_map[layer]
    url = f"https://api.vworld.kr/req/data?service=data&request=GetFeature&data={v_layer}&key={vworld_key}&domain=http://127.0.0.1&size=1000&geometry=false"
    
    if parent_code:
        url += f"&attrFilter={code_field}:like:{parent_code}"
        
    try:
        res = requests.get(url, timeout=10).json()
        if res.get('response', {}).get('status') == 'OK':
            features = res['response']['result']['featureCollection']['features']
            
            data_list = []
            for f in features:
                props = f['properties']
                data_list.append({
                    "code": props[code_field],
                    "name": props.get(name_field, props[code_field])
                })
            
            data_list.sort(key=lambda x: x['name'])
            return jsonify({"success": True, "data": data_list})
            
        return jsonify({"success": True, "data": []})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/law_review')
def law_review():
    return render_template('law_review.html')



@app.route('/api/supervisor/checklist', methods=['GET', 'POST'])
def get_supervisor_checklist():
    try:
        file_path = os.path.join(os.path.dirname(__file__), 'supervisor_db.json')
        with open(file_path, 'r', encoding='utf-8') as f:
            db_data = json.load(f)
            
        if request.method == 'POST' and request.json:
            project_data = request.json
            if project_data.get('description') or project_data.get('budget'):
                if GEMINI_KEY:
                    genai.configure(api_key=GEMINI_KEY)
                    try:
                        model = genai.GenerativeModel('models/gemini-2.5-flash')
                    except:
                        model = genai.GenerativeModel('models/gemini-2.0-flash')
                    
                    prompt = f"""
당신은 현장 공사감독관을 위한 맞춤형 체크리스트 필터링 AI입니다.
아래 [사업 개요]를 꼼꼼히 읽고, 이어지는 [전체 체크리스트 DB]의 항목(task) 중에서 이 공사에 **해당하지 않거나 불필요한 항목의 task_id**를 추출하세요.

[사업 개요]
- 사업명: {project_data.get('projectName', '')}
- 예산: {project_data.get('budget', 0)}억 원
- 면적: {project_data.get('totalArea', 0)}㎡
- 주요 사업 내용: {project_data.get('description', '')}

[판단 기준 예시]
1. 건설사업관리(감리) 항목: "직접 감독" 공사이거나 소규모 공사인 경우 감리가 없을 수 있습니다. 단, 사업 내용에 명확히 없다고 하지 않으면 일단 둡니다.
2. 지하안전평가: 지하 10m 이상 굴착이나 흙막이가 명시되지 않은 단순 지상/포장 공사면 제외.
3. 건축허가/건축물 사용승인: 건축물이 포함되지 않은 순수 토목(도로, 공원, 하천 정비 등)이면 제외.
4. 특정 공사 규모에 미달: 100억 이상일 때만 하는 VE(설계경제성검토) 등. 예산이 기준 미달이면 제외.
(그 외에도 공사 내용과 전혀 무관한 항목은 과감히 제외하여 실무자의 피로도를 낮추세요.)

[전체 체크리스트 DB]
{json.dumps(db_data, ensure_ascii=False)}

응답은 오직 제외할 항목의 task_id들만 순수한 JSON 배열 포맷(예: ["TSK_STG_PRE_001", "TSK_STG_CON_005"])으로 반환하세요. 마크다운(` ``` `) 없이 배열만 반환하세요.
제외할 항목이 없으면 빈 배열 []을 반환하세요.
"""
                    try:
                        response = model.generate_content(prompt)
                        resp_text = response.text.strip()
                        if resp_text.startswith("```json"): resp_text = resp_text[7:]
                        if resp_text.startswith("```"): resp_text = resp_text[3:]
                        if resp_text.endswith("```"): resp_text = resp_text[:-3]
                        
                        excluded_tasks = json.loads(resp_text.strip())
                        if isinstance(excluded_tasks, list):
                            for stage in db_data.get("project_stages", []):
                                original_checklists = stage.get("checklists", [])
                                stage["checklists"] = [task for task in original_checklists if task.get("task_id") not in excluded_tasks]
                    except Exception as e:
                        print("AI filtering error:", e)
                        pass

        return jsonify({"success": True, "data": db_data})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/report')
def report():
    return render_template('report.html')

@app.route('/other_review')
def other_review():
    return render_template('other_review.html')

@app.route('/other_review_result')
def other_review_result():
    # 검토 완료 시 결과를 새 탭으로 띄우기 위한 전용 페이지 (요청: "AI검토 완료되면
    # 새로운 페이지에 뜨는걸로 바꿔줘"). 첨부파일을 다시 참조하는 전용 채팅
    # (/api/chat/other_review)을 그대로 쓰려고 report.html 공용 페이지 대신 분리함.
    return render_template('other_review_result.html')

@app.route('/duty_list')
def duty_list():
    return render_template('duty_list.html')

@app.route('/bid_predictor')
def bid_predictor():
    return render_template('bid_predictor.html')

@app.route('/design_review')
def design_review():
    return render_template('design_review.html')

@app.route('/commencement_review')
def commencement_review():
    return render_template('commencement_review.html')

@app.route('/api/extract_parcel_from_drawing', methods=['POST'])
def extract_parcel_from_drawing():
    try:
        if 'file' not in request.files:
            return jsonify({"success": False, "message": "파일이 없습니다."}), 400
            
        file = request.files['file']
        if not file.filename:
            return jsonify({"success": False, "message": "선택된 파일이 없습니다."}), 400
            
        if not GEMINI_KEY:
            return jsonify({"success": False, "message": "Gemini API 키가 설정되지 않았습니다."}), 500

        import uuid
        temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        ext = os.path.splitext(file.filename)[1]
        temp_path = os.path.join(temp_dir, f"drawing_{uuid.uuid4().hex}{ext}")
        file.save(temp_path)
        
        genai.configure(api_key=GEMINI_KEY)
        
        try:
            uploaded_file = upload_file_for_gemini(temp_path)

            prompt = """
            당신은 도면, 지적도, 사업계획서에서 편입 부지 목록을 정확하게 추출하는 AI입니다.
            첨부된 이미지 또는 PDF에서 프로젝트에 편입되는 대상 부지의 '주소(지번)'와 '면적(㎡)' 데이터를 모두 찾아내어 아래 JSON 배열 형식으로 반환해주세요.
            순수 JSON 데이터만 반환해야 하며, 마크다운 코드블럭(```json)이나 다른 설명은 절대 추가하지 마세요.
            만약 면적이 적혀있지 않다면 "" (빈 문자열)로 두세요. 주소는 식별 가능한 최대한(시군구 포함) 적어주세요.
            
            반환 형식 예시:
            [
              {"address": "서울시 강남구 역삼동 123-4", "area": "500"},
              {"address": "경남 남해군 상주면 양아리 산 12", "area": ""}
            ]
            """
            
            try:
                available_models = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods]
            except:
                available_models = []
            model_name = None
            for preferred in ['models/gemini-2.5-flash', 'models/gemini-2.0-flash', 'models/gemini-1.5-flash']:
                if preferred in available_models:
                    model_name = preferred
                    break
            if not model_name:
                model_name = 'models/gemini-2.5-flash'

            model = genai.GenerativeModel(model_name)
            response = model.generate_content([prompt, gemini_file_part(uploaded_file)])

            resp_text = response.text.strip()
            if resp_text.startswith("```json"): resp_text = resp_text[7:]
            if resp_text.startswith("```"): resp_text = resp_text[3:]
            if resp_text.endswith("```"): resp_text = resp_text[:-3]
            resp_text = resp_text.strip()
            
            try:
                parsed_data = json.loads(resp_text)
                if not isinstance(parsed_data, list):
                    parsed_data = []
            except:
                parsed_data = []
                
        finally:
            try:
                os.remove(temp_path)
            except:
                pass
                
        return jsonify({"success": True, "data": parsed_data})
        
    except Exception as e:
        print(f"Extract Drawing Error: {e}")
        return jsonify({"success": False, "message": f"서버 오류: {str(e)}"}), 500

@app.route('/api/search_law_list', methods=['POST'])
def search_law_list():
    try:
        data = request.get_json(silent=True) or {}
        keyword = data.get('keyword', '')
        if isinstance(keyword, str):
            keyword = keyword.strip()
            
        if not keyword:
            return jsonify({"success": False, "message": "검색어를 입력해주세요."})
        
        search_res = requests.get(f"https://www.law.go.kr/DRF/lawSearch.do?OC={LAW_KEY}&target=law&type=XML&query={urllib.parse.quote(keyword)}", timeout=5)
        root = ET.fromstring(search_res.text)
        laws = []
        for law in root.findall('law'):
            laws.append({
                'lsi_seq': law.findtext('법령일련번호'),
                'law_name': law.findtext('법령명한글')
            })
        return jsonify({"success": True, "data": laws})
    except Exception as e:
        print("Law List Search Error:", e)
        return jsonify({"success": False, "message": f"법령 목록 검색 오류: {str(e)}"})


LAW_TEXT_CACHE = {}

@app.route('/api/search_duties_chunk', methods=['POST'])
def search_duties_chunk():
    try:
        data = request.get_json(silent=True) or {}
        lsi_seq = data.get('lsi_seq', '')
        exact_law_name = data.get('law_name', '')
        chunk_index = int(data.get('chunk_index', 0))
        
        if isinstance(lsi_seq, str): lsi_seq = lsi_seq.strip()
        if isinstance(exact_law_name, str): exact_law_name = exact_law_name.strip()
        
        if not lsi_seq or not exact_law_name:
            return jsonify({"success": False, "message": "법령일련번호 또는 법률명이 누락되었습니다."})
        
        # 1. Fetch XML (with memory cache to prevent rate limits)
        global LAW_TEXT_CACHE
        full_text = LAW_TEXT_CACHE.get(lsi_seq)
        
        if not full_text:
            doc_res = requests.get(f"https://www.law.go.kr/DRF/lawService.do?OC={LAW_KEY}&target=law&type=XML&MST={lsi_seq}", timeout=10)
            doc_root = ET.fromstring(doc_res.text)
            
            articles = doc_root.findall('.//조문단위')
            full_text = ""
            for art in articles:
                art_title = art.findtext('조문내용') or ""
                full_text += art_title + "\\n"
                for hang in art.findall('.//항내용'):
                    full_text += hang.text + "\\n"
                for ho in art.findall('.//호내용'):
                    full_text += ho.text + "\\n"
            
            LAW_TEXT_CACHE[lsi_seq] = full_text
                
        # 2. Split into chunks (1,500 chars per chunk to ensure < 40s even with detailed JSON)
        CHUNK_SIZE = 1500
        chunks = [full_text[i:i+CHUNK_SIZE] for i in range(0, len(full_text), CHUNK_SIZE)]
        if not chunks:
            chunks = [""]
            
        if chunk_index >= len(chunks):
            return jsonify({"success": True, "data": {"duties": [], "has_more": False, "total_chunks": len(chunks)}})
            
        current_chunk = chunks[chunk_index]
            
        if not GEMINI_KEY:
            return jsonify({"success": False, "message": "Gemini API 키가 설정되지 않았습니다."})
            
        genai.configure(api_key=GEMINI_KEY)
        model_name = 'models/gemini-2.5-flash'
        model = genai.GenerativeModel(model_name)
        
        # 4. Prompt without 7-item limit but allowing detailed output
        prompt = f"""
다음은 '{exact_law_name}' 법령의 일부 조문입니다 (파트 {chunk_index + 1}/{len(chunks)}):
{current_chunk}

이 법령 내용 중에서 '행정/건설 관리기관, 사업주, 지자체 등이 의무적으로 이행해야 하는 사항'(예: 정기 안전점검, 교육 실시, 계획 수립, 결과 통보 등)만 모두 추출하세요.
(발견되는 모든 의무사항을 남김없이 전부 추출하되, 각 의무의 내용을 충분히 구체적이고 상세하게 설명하세요.)

🚨 [조문번호 관련 필수 규칙]: "article" 값은 위 조문 텍스트에 실제로 적힌 조 번호(예: "제3조", "제5조제3항")를 그대로 옮기십시오.
아래 JSON 형식 예시의 "제O조"는 형식을 보여주기 위한 자리표시자일 뿐, 실제 조문 번호가 아닙니다.
텍스트에서 해당 의무가 몇 조에 있는지 정확히 찾을 수 없으면 "제O조"라고 쓰지 말고 "조 번호 확인 필요"라고 쓰십시오.

결과는 오직 아래의 순수 JSON 배열 포맷으로만 반환하세요(마크다운 없이). 의무사항이 없으면 빈 배열 []을 반환하세요.
[
  {{
    "article": "(실제 조 번호, 예: 제3조)",
    "duty_title": "핵심 의무 제목",
    "description": "구체적인 의무 내용을 상세하게 서술 (이행해야 할 대상, 조건, 방법 등을 포함)",
    "frequency": "수시 / 연 1회 등 기한",
    "target": "의무 이행 주체"
  }}
]
"""
        response = model.generate_content(prompt)
        resp_text = response.text.strip()
        if resp_text.startswith("```json"): resp_text = resp_text[7:]
        if resp_text.startswith("```"): resp_text = resp_text[3:]
        if resp_text.endswith("```"): resp_text = resp_text[:-3]
        
        try:
            duties = json.loads(resp_text.strip())
        except json.JSONDecodeError:
            print("JSON Decode Error. Raw resp:", resp_text)
            duties = []
            
        result_data = {
            "law_name": exact_law_name,
            "duties": duties,
            "has_more": chunk_index < len(chunks) - 1,
            "total_chunks": len(chunks)
        }
        
        return jsonify({"success": True, "data": result_data})

    except Exception as e:
        print("Chunked Duty Search Error:", e)
        return jsonify({"success": False, "message": f"서버 오류: {str(e)}"})


def _build_bid_histogram(rates, bin_width=0.2):
    """낙찰률 목록을 bin_width(%p) 단위 구간으로 나눠 차트용 히스토그램 데이터를 만든다."""
    if not rates:
        return []
    lo = (min(rates) // bin_width) * bin_width
    hi = max(rates)
    buckets = {}
    n_bins = int((hi - lo) / bin_width) + 2
    for r in rates:
        idx = int((r - lo) / bin_width)
        buckets[idx] = buckets.get(idx, 0) + 1
    result = []
    for idx in range(n_bins):
        start = round(lo + idx * bin_width, 2)
        result.append({"range": start, "count": buckets.get(idx, 0)})
    return result


@app.route('/api/bid_agency_search', methods=['GET'])
def api_bid_agency_search():
    """발주기관명 입력창 자동완성. 나라장터 최근 낙찰 데이터에 실제 등장한 기관명
    중에서 검색어를 포함하는 것들을 찾아 후보로 준다(전용 기관명 검색 API가 없어서
    실 데이터로 목록을 직접 구축 - pps_bid.search_agency_names 참고)."""
    try:
        from pps_bid import search_agency_names
        query = request.args.get('q', '')
        return jsonify({"success": True, "data": search_agency_names(query)})
    except Exception as e:
        return jsonify({"success": False, "message": str(e), "data": []})


@app.route('/api/open_bids', methods=['GET'])
def api_open_bids():
    """'진행 중인 입찰공고 목록' - 아직 입찰마감이 지나지 않은 공고를 검색해서 보여준다.
    여기서 공고를 하나 고르면(기초금액/발주기관을 그대로 가져다) 바로 낙찰가 예측으로
    이어갈 수 있게 하기 위함(요청: "입찰해야 하는 정보도 보여주고 싶은거지... 기초금액
    보고 어디 발주처인지 보고 낙찰가를 예견하는거지")."""
    try:
        from pps_bid import fetch_open_bids

        dminstt_nm = (request.args.get('dminsttNm') or '').strip() or None
        keyword = (request.args.get('keyword') or '').strip() or None
        region = (request.args.get('region') or '').strip() or None

        items, err = fetch_open_bids(days=30, dminstt_nm=dminstt_nm, keyword=keyword, region=region)
        if err:
            return jsonify({"success": False, "message": err, "data": []})

        # 목록 카드 렌더링에 필요한 필드만 추려서 반환
        data = [{
            "bidNtceNo": it.get("bidNtceNo"),
            "bidNtceOrd": it.get("bidNtceOrd"),
            "bidNtceNm": it.get("bidNtceNm"),
            "dminsttNm": it.get("dminsttNm"),
            "ntceInsttNm": it.get("ntceInsttNm"),
            "bidClseDt": it.get("bidClseDt"),
            "opengDt": it.get("opengDt"),
            "presmptPrce": it.get("presmptPrce"),
            "bdgtAmt": it.get("bdgtAmt"),
            "cnstrtsiteRgnNm": it.get("cnstrtsiteRgnNm"),
            "sucsfbidMthdNm": it.get("sucsfbidMthdNm"),
            "mainCnsttyNm": it.get("mainCnsttyNm"),
            "bidNtceDtlUrl": it.get("bidNtceDtlUrl"),
        } for it in items]

        return jsonify({"success": True, "data": data, "count": len(data)})
    except Exception as e:
        return jsonify({"success": False, "message": str(e), "data": []})


@app.route('/api/open_bids/base_amount', methods=['GET'])
def api_open_bids_base_amount():
    """특정 공고의 정확한 기초금액(bssamt)을 조회한다. 일부 공고(용역성/소액 등)는
    기초금액 자료가 아직 없을 수 있어 그 경우 실패를 명시적으로 알린다(프론트에서는
    이미 목록에 있는 추정가격/예산금액으로 대체 표시)."""
    try:
        from pps_bid import fetch_bid_base_amount

        bid_ntce_no = request.args.get('bidNtceNo', '')
        bid_ntce_ord = request.args.get('bidNtceOrd', '000')

        match, err = fetch_bid_base_amount(bid_ntce_no, bid_ntce_ord)
        if err:
            return jsonify({"success": False, "message": err})

        return jsonify({
            "success": True,
            "bssamt": match.get("bssamt"),
            "rsrvtnPrceRngBgnRate": match.get("rsrvtnPrceRngBgnRate"),
            "rsrvtnPrceRngEndRate": match.get("rsrvtnPrceRngEndRate"),
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route('/api/analyze/bid_predict', methods=['POST'])
def api_bid_predict():
    """'낙찰가 예측' - 조달청 나라장터 과거 공사 낙찰 이력을 통계 내어, 이번 입찰에서
    경쟁력 있는 낙찰률(%) 구간을 확률적으로 가늠할 수 있게 돕는다.

    🚨 주의: 적격심사 방식은 낙찰 전에 예정가격 자체가 복수예비가격 추첨으로 정해지므로,
    특정 금액을 "이거다"라고 확정 예측하는 것은 원리상 불가능하다. 이 기능은 "비슷한
    발주기관/공종/규모의 과거 낙찰들이 실제로 어느 낙찰률에 몰려 있었는지"를 실제
    공공데이터로 통계 낼 뿐, AI가 숫자를 만들어내지 않는다(전부 파이썬 통계 계산).
    """
    try:
        from pps_bid import fetch_bid_history, compute_rate_stats, percentile_of_value

        data = request.json or {}
        dminstt_nm = (data.get('dminsttNm') or '').strip() or None
        keyword = (data.get('keyword') or '').strip() or None
        region = (data.get('region') or '').strip() or None

        try:
            months = int(data.get('months', 12))
        except (TypeError, ValueError):
            months = 12
        months = max(1, min(months, 24))

        price_min = price_max = None
        presumed_price = data.get('presumedPrice')
        if presumed_price:
            try:
                p = float(presumed_price)
                if p > 0:
                    price_min, price_max = p * 0.5, p * 2.0
            except (TypeError, ValueError):
                pass

        if not dminstt_nm and not keyword and not region:
            return jsonify({"success": False, "message": "발주기관명, 공고 키워드, 참가제한지역 중 최소 하나는 입력해 주세요. (전체 조회는 데이터가 너무 많아 지원하지 않습니다.)"}), 400

        items, errors = fetch_bid_history(
            months=months, dminstt_nm=dminstt_nm, keyword=keyword, region=region,
            presumed_price_min=price_min, presumed_price_max=price_max,
        )

        stats = compute_rate_stats(items)
        if not stats:
            return jsonify({
                "success": False,
                "message": "조건에 맞는 낙찰 이력을 찾을 수 없습니다. 조회 기간을 늘리거나 검색 조건(발주기관/키워드/지역)을 넓혀서 다시 시도해 주세요.",
                "errors": errors,
            })

        rates = stats.pop('raw_rates')

        target_percentile = None
        target_rate = data.get('targetRate')
        if target_rate not in (None, ''):
            try:
                target_percentile = percentile_of_value(rates, float(target_rate))
            except (TypeError, ValueError):
                pass

        # 기초금액을 입력하면 각 백분위(%)를 실제 목표 입찰가(원)로 환산해서 보여준다
        # (요청: "저걸 어떻게 사용하면 낙찰가를 확률적으로 근접할 수 있을까" ->
        # %만으로는 실전에서 바로 못 쓰니 원 단위 금액까지 계산해달라는 후속 요청).
        # 🚨 주의: sucsfbidRate(낙찰률)는 "예정가격 대비" 비율인데, 예정가격은 복수예비
        # 가격 추첨으로 기초금액에서 통상 ±2~3% 정도 벗어날 수 있어 정확히 알 수 없다.
        # 그래서 "기초금액 x 낙찰률"은 근사 추정치일 뿐이며, 이 오차 범위(±3%)를 함께
        # 보여줘서 단일 숫자를 과신하지 않도록 한다.
        won_estimates = None
        base_amount = data.get('baseAmount')
        if base_amount not in (None, ''):
            try:
                base = float(base_amount)
                if base > 0:
                    PRICE_VOLATILITY = 0.03  # 예정가격이 기초금액 대비 벌어질 수 있는 대략적 범위
                    pct_keys = ['min', 'p10', 'p25', 'p40', 'p50', 'p60', 'p75', 'p90', 'max']
                    won_estimates = {
                        "base_amount": base,
                        "volatility": PRICE_VOLATILITY,
                        "by_percentile": {
                            key: {
                                "rate": stats[key],
                                "amount": round(base * stats[key] / 100),
                                "amount_low": round(base * (1 - PRICE_VOLATILITY) * stats[key] / 100),
                                "amount_high": round(base * (1 + PRICE_VOLATILITY) * stats[key] / 100),
                            }
                            for key in pct_keys
                        },
                    }
            except (TypeError, ValueError):
                pass

        return jsonify({
            "success": True,
            "stats": stats,
            "histogram": _build_bid_histogram(rates),
            "sample_size": len(items),
            "used_filters": {"dminsttNm": dminstt_nm, "keyword": keyword, "region": region, "months": months},
            "target_rate": target_rate,
            "target_percentile": target_percentile,
            "won_estimates": won_estimates,
            "errors": errors,
        })
    except Exception as e:
        print("Bid Predict Error:", e)
        return jsonify({"success": False, "message": f"서버 오류: {str(e)}"}), 500


@app.route('/api/analyze/bid_backtest', methods=['POST'])
def api_bid_backtest():
    """'전략 백테스트' - 조회 기간을 학습 구간(예전)/검증 구간(최근)으로 나눠서,
    학습 구간에서 뽑은 추천 낙찰률(백분위)을 검증 구간의 실제 낙찰 결과에 그대로
    대입했다면 몇 %나 이겼을지 사후 검증한다. "낙찰가 예측" 화면의 추천이 순환논리에
    그치지 않고 시간이 지나도 실제로 통하는지 확인하기 위한 기능(요청: "백테스트 할 수
    있는 기능도 만들자")."""
    try:
        from pps_bid import backtest_bid_strategy

        data = request.json or {}
        dminstt_nm = (data.get('dminsttNm') or '').strip() or None
        keyword = (data.get('keyword') or '').strip() or None
        region = (data.get('region') or '').strip() or None

        try:
            months = int(data.get('months', 12))
        except (TypeError, ValueError):
            months = 12
        months = max(2, min(months, 24))

        try:
            test_months = int(data.get('testMonths', 3))
        except (TypeError, ValueError):
            test_months = 3
        test_months = max(1, min(test_months, months - 1))

        price_min = price_max = None
        presumed_price = data.get('presumedPrice')
        if presumed_price:
            try:
                p = float(presumed_price)
                if p > 0:
                    price_min, price_max = p * 0.5, p * 2.0
            except (TypeError, ValueError):
                pass

        if not dminstt_nm and not keyword and not region:
            return jsonify({"success": False, "message": "발주기관명, 공고 키워드, 참가제한지역 중 최소 하나는 입력해 주세요."}), 400

        result, err = backtest_bid_strategy(
            months=months, test_months=test_months, dminstt_nm=dminstt_nm, keyword=keyword,
            region=region, presumed_price_min=price_min, presumed_price_max=price_max,
        )
        if err:
            return jsonify({"success": False, "message": err})

        return jsonify({"success": True, **result})
    except Exception as e:
        print("Bid Backtest Error:", e)
        return jsonify({"success": False, "message": f"서버 오류: {str(e)}"}), 500


@app.route('/api/search_duties', methods=['POST'])
def search_duties():
    try:
        data = request.get_json(silent=True) or {}
        lsi_seq = data.get('lsi_seq', '')
        exact_law_name = data.get('law_name', '')
        if isinstance(lsi_seq, str): lsi_seq = lsi_seq.strip()
        if isinstance(exact_law_name, str): exact_law_name = exact_law_name.strip()
        
        if not lsi_seq or not exact_law_name:
            return jsonify({"success": False, "message": "법령일련번호 또는 법률명이 누락되었습니다."})
        
        doc_res = requests.get(f"https://www.law.go.kr/DRF/lawService.do?OC={LAW_KEY}&target=law&type=XML&MST={lsi_seq}", timeout=10)
        doc_root = ET.fromstring(doc_res.text)
        
        articles = doc_root.findall('.//조문단위')
        full_text = ""
        for art in articles:
            art_title = art.findtext('조문내용') or ""
            full_text += art_title + "\n"
            for hang in art.findall('.//항내용'):
                full_text += hang.text + "\n"
            for ho in art.findall('.//호내용'):
                full_text += ho.text + "\n"
                
        if len(full_text) > 8000:
            full_text = full_text[:8000]
            
        if not GEMINI_KEY:
            return jsonify({"success": False, "message": "Gemini API 키가 설정되지 않았습니다."})
            
        genai.configure(api_key=GEMINI_KEY)
        model_name = 'models/gemini-2.5-flash'
    
        model = genai.GenerativeModel(model_name)
        prompt = f"""
다음은 '{exact_law_name}' 법령의 일부 조문입니다:
{full_text}

이 법령 내용 중에서 '행정/건설 관리기관, 사업주, 지자체 등이 의무적으로 이행해야 하는 사항'(예: 정기 안전점검, 교육 실시, 계획 수립, 결과 통보 등)만 추출하세요.
(응답 시간 최적화를 위해 가장 중요하고 핵심적인 의무 사항 최대 7개까지만 추출하세요.)
결과는 오직 아래의 순수 JSON 배열 포맷으로만 반환하세요(마크다운 없이). 의무사항이 없으면 빈 배열 []을 반환하세요.
[
  {{
    "article": "제O조",
    "duty_title": "핵심 의무 제목",
    "description": "구체적인 의무 내용 요약",
    "frequency": "수시 / 연 1회 등 기한",
    "target": "의무 이행 주체"
  }}
]
"""
        response = model.generate_content(prompt)
        resp_text = response.text.strip()
        if resp_text.startswith("```json"): resp_text = resp_text[7:]
        if resp_text.startswith("```"): resp_text = resp_text[3:]
        if resp_text.endswith("```"): resp_text = resp_text[:-3]
        
        duties = json.loads(resp_text.strip())
        
        result_data = {
            "law_name": exact_law_name,
            "duties": duties
        }
        
        return jsonify({"success": True, "data": result_data})
        
    except Exception as e:
        print("Duty Search Error:", e)
        return jsonify({"success": False, "message": f"서버 오류: {str(e)}"})

@app.route('/api/debug')
def debug_env():
    key = os.environ.get('GEMINI_API_KEY', '')
    masked_key = key[:10] + "..." + key[-5:] if len(key) > 15 else "EMPTY_OR_TOO_SHORT"
    return jsonify({
        "gemini_key_in_server": masked_key,
        "message": "서버에 현재 등록된 API 키의 앞/뒷부분입니다. 발급받으신 새 키와 일치하는지 확인해주세요."
    })
def fetch_law_data(law_key, search_query="국토의 계획 및 이용에 관한 법률"):
    if not law_key:
        return "법제처 API 키가 제공되지 않아 AI 자체 지식을 기반으로 분석합니다."
    try:
        url = f"https://www.law.go.kr/DRF/lawSearch.do?OC={law_key}&target=law&type=XML&query={search_query}"
        response = requests.get(url, timeout=5)
        response.raise_for_status()
        import xml.etree.ElementTree as ET
        root = ET.fromstring(response.content)
        laws = []
        for law in root.findall('.//law'):
            law_name = law.find('법령명한글').text if law.find('법령명한글') is not None else ""
            laws.append(law_name)
        if laws:
            return f"법제처 API에서 조회된 연관 법령 목록: {', '.join(laws[:5])} 등."
        else:
            return "연관 법령을 찾지 못했습니다."
    except Exception as e:
        return "법제처 API 통신 중 오류가 발생하여 자체 지식을 활용합니다."

def download_law_to_db(law_name, law_key, md_path):
    import os
    import xml.etree.ElementTree as ET
    import urllib.parse
    import requests
    try:
        url = f'https://www.law.go.kr/DRF/lawSearch.do?OC={law_key}&target=law&type=XML&query={urllib.parse.quote(law_name)}'
        res = requests.get(url, timeout=5)
        root = ET.fromstring(res.text)
        law = root.find('.//law')
        if law is None:
            return False
        lsi_seq = law.find('법령일련번호').text
        
        doc_url = f'https://www.law.go.kr/DRF/lawService.do?OC={law_key}&target=law&type=XML&MST={lsi_seq}'
        doc_res = requests.get(doc_url, timeout=10)
        doc_root = ET.fromstring(doc_res.text)
        
        with open(md_path, 'w', encoding='utf-8') as f:
            f.write(f'# {law_name}\n\n')
            for art in doc_root.findall('.//조문단위'):
                num = art.findtext('조문번호')
                title = art.findtext('조문제목')
                body = art.findtext('조문내용')
                title_str = f'({title})' if title else ''
                f.write(f'### 제{num}조 {title_str}\n{body}\n\n')
                
                for hang in art.findall('.//항'):
                    hang_body = hang.findtext('항내용')
                    if hang_body: f.write(f'{hang_body}\n')
                    for ho in hang.findall('.//호'):
                        ho_body = ho.findtext('호내용')
                        if ho_body: f.write(f'  {ho_body}\n')
                    for mok in hang.findall('.//목'):
                        mok_body = mok.findtext('목내용')
                        if mok_body: f.write(f'    {mok_body}\n')
                f.write('\n')
        return True
    except Exception as e:
        print(f"Auto-download failed for {law_name}: {e}")
        return False

def extract_keyword_via_llm(text):
    try:
        genai.configure(api_key=GEMINI_KEY)
        kw_prompt = f"다음 텍스트에서 대한민국 법제처 판례/법령 검색에 가장 적합한 핵심 명사 키워드 딱 1개(예: 영업손실보상, 하도급, 직불)만 추출해. 다른 말은 절대 하지마.\n텍스트: {text[:3000]}"
        
        try:
            available_models = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods and 'vision' not in m.name.lower()]
            models_to_try = sorted(available_models, key=lambda x: (0 if '1.5-pro' in x else 1 if '2.5-pro' in x else 2 if 'pro' in x else 3 if '1.5-flash' in x else 4))
        except:
            models_to_try = ['models/gemini-2.5-flash', 'models/gemini-2.0-flash', 'models/gemini-1.5-flash']
            
        for m in models_to_try:
            try:
                model = genai.GenerativeModel(m)
                kw_res = model.generate_content(kw_prompt)
                keyword = kw_res.text.strip().replace("'", "").replace('"', "")
                if len(keyword) > 10: keyword = keyword[:10]
                if keyword: return keyword
            except Exception as e:
                continue
    except Exception as e:
        pass
    return ""


def extract_cases_via_llm(text, uploaded_file=None):
    try:
        genai.configure(api_key=GEMINI_KEY)
        kw_prompt = "다음 텍스트나 첨부된 문서에서 인용된 '대법원 판례 사건번호(예: 2010두11641)'를 모두 찾아내어 쉼표로 구분해 줘. 판례가 없으면 '없음'이라고 해.\n텍스트: " + text[:3000]
        
        contents_payload = [kw_prompt]
        if uploaded_file:
            contents_payload.append(uploaded_file)
            
        try:
            available_models = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods and 'vision' not in m.name.lower()]
            models_to_try = sorted(available_models, key=lambda x: (0 if '1.5-flash' in x else 1 if '1.5-pro' in x else 2 if '2.5-flash' in x else 3 if 'pro' in x else 4))
        except:
            models_to_try = ['models/gemini-1.5-flash', 'models/gemini-2.5-flash', 'models/gemini-2.0-flash']
            
        for m in models_to_try:
            try:
                model = genai.GenerativeModel(m)
                res = model.generate_content(contents_payload)
                result_text = res.text.strip().replace(" ", "").replace("'", "").replace('"', "")
                if '없음' not in result_text and len(result_text) > 4:
                    return result_text
                return ""
            except Exception as e:
                continue
    except Exception as e:
        pass
    return ""


def fetch_local_law_data(query, moleg_context):
    import glob
    import os
    import re
    local_data = ""
    # Check data/laws directory
    laws_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'laws')
    if not os.path.exists(laws_dir):
        return ""

    # 질의(+MCP 컨텍스트)에서 의미 있는 키워드(3자 이상)를 뽑아, 그 키워드가
    # 파일명에 포함되는지로 매칭한다.
    # (예전에는 반대로 "파일명 전체가 질의 문장 안에 통째로 들어있는지"를 확인했는데,
    #  사용자는 "재해영향평가"처럼 짧게 묻고 파일명은 "재해영향평가등의_협의_실무지침"처럼
    #  훨씬 길어서 거의 매칭되지 않았다. 그 결과 관련 로컬 법령이 전혀 검색되지 않고
    #  AI가 사전지식만으로 답하다가 완전히 다른 법(환경영향평가법)을 재해영향평가로
    #  착각해 답하는 사고로 이어짐.)
    combined_text = f"{query} {moleg_context}"
    keywords = set(w for w in re.split(r'[\s,./()·\[\]]+', combined_text) if len(w) >= 3)

    # 위 키워드 매칭은 "질의 단어가 파일명에 연속해서 그대로 들어있는지"만 보기 때문에,
    # 정식 법령명이 길어서 실무에서 줄임말(약칭)로 더 자주 불리는 법은 여전히 못 찾는다.
    # 예: "토지보상법"은 정식명 "공익사업을_위한_토지_등의_취득_및_보상에_관한_법률" 안에
    # "토지"와 "보상"이 붙어있지 않아 통짜 키워드로는 매칭되지 않고, 로컬 DB에 원문이
    # 있는데도 "원문 없음"으로 잘못 안내되는 사고로 이어짐(실사용 중 발견).
    # 같은 패턴의 법들에 대해 약칭 -> 정식명 키워드 별칭을 추가해준다.
    LAW_ABBREVIATIONS = {
        '토지보상법': '공익사업을위한토지등의취득및보상에관한법률',
        '국토계획법': '국토의계획및이용에관한법률',
        '지방계약법': '지방자치단체를당사자로하는계약에관한법률',
        '건진법': '건설기술진흥법',
        '건산법': '건설산업기본법',
        '시설물안전법': '시설물의안전및유지관리에관한특별법',
        '지하안전법': '지하안전관리에관한특별법',
        '재난안전법': '재난및안전관리기본법',
        '건설폐기물법': '건설폐기물의재활용촉진에관한법률',
    }
    # 실제 질의에서는 "토지보상법에", "국토계획법상" 처럼 조사가 붙어 나오는 경우가
    # 대부분이라 완전 일치(kw == 약칭)로는 걸리지 않는다. 약칭이 키워드 안에
    # 부분 포함되는지로 확인한다.
    for kw in list(keywords):
        for abbr, full_name in LAW_ABBREVIATIONS.items():
            if abbr in kw:
                keywords.add(full_name)

    # 파일명 매칭용 keywords는 오탐 방지를 위해 3자 이상만 쓰지만, "인도"/"철거"처럼
    # 정작 조문 안에서는 2자 단어가 핵심인 경우가 많다. 파일 "안"에서 관련 부분을
    # 찾을 때는 원 질의(모호한 MCP 검색결과 산문은 제외)에서 2자 이상까지 넓게 뽑는다.
    locate_keywords = keywords | set(w for w in re.split(r'[\s,./()·\[\]]+', query) if len(w) >= 2)

    # 파일 하나가 통째로(최대 1.8MB짜리도 있음) 프롬프트를 잡아먹지 않도록 파일당/전체 상한을 둔다.
    # (moleg_context는 실시간 검색 결과 산문이라 법령명을 여러 개 언급하기 쉬워서,
    #  키워드 매칭 정확도를 올리자마자 거대 파일이 여러 개 한꺼번에 걸려 프롬프트가
    #  비정상적으로 커지고 응답이 몇 분씩 걸리는 사고로 실제 이어졌음.)
    MAX_PER_FILE = 8000
    MAX_TOTAL = 24000

    for md_file in glob.glob(os.path.join(laws_dir, '*.md')):
        if len(local_data) >= MAX_TOTAL:
            break
        law_name_key = os.path.basename(md_file).replace('.md', '').replace('_', '')

        if any(kw in law_name_key for kw in keywords):
            try:
                with open(md_file, "r", encoding="utf-8") as f:
                    content = f.read()
                remaining = MAX_TOTAL - len(local_data)
                local_data += _extract_relevant_excerpt(content, locate_keywords, min(MAX_PER_FILE, remaining)) + "\n\n"
            except:
                pass
    return local_data


def _extract_relevant_excerpt(content, keywords, max_len):
    """content가 max_len보다 길면 앞부분만 자르지 않고, 키워드가 실제로 등장하는
    지점 주변을 발췌한다.

    (예전에는 그냥 content[:max_len]으로 앞부분만 잘랐는데, 토지보상법처럼
    조문이 많은(5만자 이상) 법은 정작 물어본 조항(예: 제43조 토지 인도)이
    파일 뒷부분에 있어서 통째로 잘려나가고, 로컬 DB에 원문이 있는데도
    "원문 없음"이라고 잘못 안내하는 사고로 이어졌음 - 실사용 중 발견.)
    """
    import re
    if len(content) <= max_len:
        return content

    # 키워드별로 문서 안에서 몇 번이나 등장하는지 센다. "공익사업"/"보상금"처럼
    # 법 전체에 수십 번 나오는 흔한 단어는 조문 위치를 특정하는 데 도움이 안 되고,
    # 오히려 그 흔한 단어들이 문서 앞부분에 몰려있어서 예산을 다 잡아먹고 정작
    # 필요한 뒷부분 조항(드물게/한 번만 나오는 단어 근처)을 놓치는 사고로
    # 이어졌다(실사용 중 발견: "인도"는 8번뿐이라 걸러졌어야 했는데 3글자
    # 미만이라 애초에 키워드에서 빠졌던 문제와 겹쳐서 발생). 그래서 흔한 단어는
    # 아예 건너뛰고, 희귀한(구체적인) 키워드부터 우선 발췌한다.
    counts = {}
    for kw in keywords:
        if len(kw) < 2:
            continue
        c = content.count(kw)
        if 0 < c <= 20:
            counts[kw] = c
    if not counts:
        return content[:max_len]

    head = content[:400]
    budget = max_len - len(head)
    covered = [(0, len(head))]

    def overlaps(s, e):
        return any(s < ce and e > cs for cs, ce in covered)

    picks = []  # (start, piece)
    for kw, _cnt in sorted(counts.items(), key=lambda x: x[1]):  # 희귀한 키워드부터
        if budget <= 0:
            break
        for m in re.finditer(re.escape(kw), content):
            pos = m.start()
            start = max(0, pos - 150)
            end = min(len(content), pos + 1200)
            if overlaps(start, end):
                continue
            piece = content[start:end][:budget]
            picks.append((start, piece))
            covered.append((start, start + len(piece)))
            budget -= len(piece)
            break  # 키워드 하나당 대표 지점 1곳만

    picks.sort(key=lambda x: x[0])  # 읽기 자연스럽도록 원문 등장 순서로 재정렬
    return head + "\n...(중략)...\n" + "\n...(중략)...\n".join(p for _, p in picks)


_LAW_TEXT_FOR_VERIFY_CACHE = {}


def _verify_citation(item):
    """AI 법규 검토(law_review)가 인용한 법령 조항(law_name + article)이 로컬 법령
    원문(data/laws/*.md, 62개)에 실제로 존재하는지 대조한다.

    law_review는 구글 검색 그라운딩에 의존하는데, 보고서 하나에 조항 인용이
    10~20개씩 나오다 보니 AI가 매 항목을 실제로 검색해 검증하지 않고 자체 지식
    (사전학습 기억)만으로 채우는 경우가 있다. 그 결과 아래 두 가지 사고가
    실사용 중 발견됨:
      1) 존재하지 않거나 이미 폐지된 조항을 현행처럼 인용
         (예: "국토의 계획 및 이용에 관한 법률 제71조" - 실제로는 2006년에 삭제됨)
      2) 조항 번호 자체는 실존하지만 완전히 다른 내용의 조항을 잘못 인용
         (예: "건설산업기본법 제35조"를 "건설기술인 배치"에 인용했는데, 실제 제35조는
         "하도급대금의 직접 지급"이라 전혀 무관함)

    1)번은 기계적으로 확실히 걸러낼 수 있어 "확인필요"로 명확히 표시한다.
    2)번은 AI 재판단(의미 비교) 없이는 "틀렸다"고 자동 단정하기 어렵고, 섣불리
    단정하면 실제로는 맞는 인용까지 "확인필요"로 잘못 표시하는 역효과가 크다.
    대신 로컬 원문에서 그 조항의 실제 제목을 찾아 인용 옆에 그대로 보여줘서,
    "건설산업기본법 제35조(하도급대금의 직접 지급)"처럼 실제 제목과 이 항목의
    설명이 안 맞으면 사람이 한눈에 알아챌 수 있게 한다.
    (로컬에 원문이 없는 법은 대조 자체가 불가능해 그대로 둔다.)
    """
    import re
    law_name = (item.get('law_name') or '').strip()
    article_field = (item.get('article') or '').strip()
    if not law_name or not article_field:
        return

    if law_name in _LAW_TEXT_FOR_VERIFY_CACHE:
        text = _LAW_TEXT_FOR_VERIFY_CACHE[law_name]
    else:
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'laws', law_name.replace(' ', '_') + '.md')
        try:
            with open(path, 'r', encoding='utf-8') as f:
                text = f.read()
        except (FileNotFoundError, OSError):
            text = None
        _LAW_TEXT_FOR_VERIFY_CACHE[law_name] = text

    if text is None:
        return  # 로컬 DB에 없는 법 - 검증 불가

    tokens = re.findall(r'제\d+조(?:의\d+)?', article_field)
    if not tokens:
        return

    reason_key = 'reason' if 'reason' in item else 'desc'
    titles = []
    for article_no in tokens:
        if re.search(r'^' + re.escape(article_no) + r'\s*삭제', text, re.MULTILINE):
            item['article'] = f"{article_field} (확인필요)"
            item[reason_key] = (
                f"⚠️ [자동 검증] {article_no}는 이미 삭제된 조항으로 확인되어 인용 근거가 부정확할 수 있습니다. 법제처 원문을 직접 확인하십시오. "
                + (item.get(reason_key) or '')
            )
            return
        m = re.search(r'^' + re.escape(article_no) + r'\(([^)]*)\)', text, re.MULTILINE)
        if not m:
            item['article'] = f"{article_field} (확인필요)"
            item[reason_key] = (
                f"⚠️ [자동 검증] {article_no}가 로컬 법령 원문에서 확인되지 않았습니다. 법제처 원문을 직접 확인하십시오. "
                + (item.get(reason_key) or '')
            )
            return
        titles.append((article_no, m.group(1)))

    # 전부 실존하는 조항 - 실제 제목을 옆에 붙여서 사람이 주제 일치 여부를 스스로 판단할 수 있게 함
    if len(titles) == 1:
        item['article'] = f"{article_field}({titles[0][1]})"
    else:
        item['article'] = article_field + " (" + ", ".join(f"{no}: {t}" for no, t in titles) + ")"


@app.route('/api/verify_parcel', methods=['POST'])
def verify_parcel():
    data = request.json
    
    full_address = data.get('address')
    pnu = ""
    san = "1"
    
    import re
    
    if full_address and VWORLD_KEY:
        try:
            params = {
                "service": "search", "request": "search", "version": "2.0",
                "size": "10", "page": "1", "query": full_address,
                "type": "address", "category": "parcel", "format": "json",
                "errorformat": "json", "key": VWORLD_KEY.strip(), "domain": "http://127.0.0.1"
            }
            res_search = requests.get("https://api.vworld.kr/req/search", params=params, timeout=5).json()
            items = res_search.get('response', {}).get('result', {}).get('items', [])
            
            if not items:
                return jsonify({"success": False, "message": f"검색 API 결과 없음: {res_search}"}), 400
                
            # 정확한 주소 매칭 (읍/면/리/지번 검증)
            input_parts = re.split(r'\s+', full_address.strip())
            for item in items:
                api_addr = item.get('address', {}).get('parcel', '')
                if not api_addr:
                    continue
                # 마지막 부분(지번)이 일치해야 함
                if input_parts[-1] != api_addr.split()[-1]:
                    continue
                
                # 다른 부분들도 모두 포함되는지 확인 (경북/경상북도 등 예외 처리)
                is_match = True
                for p in input_parts[:-1]:
                    if p not in api_addr:
                        if p == '경북' and '경상북도' in api_addr: continue
                        if p == '경남' and '경상남도' in api_addr: continue
                        if p == '전북' and '전라북도' in api_addr: continue
                        if p == '전남' and '전라남도' in api_addr: continue
                        if p == '충북' and '충청북도' in api_addr: continue
                        if p == '충남' and '충청남도' in api_addr: continue
                        is_match = False
                        break
                
                if is_match:
                    pnu = item.get('id', '')
                    break
            
            if not pnu:
                return jsonify({"success": False, "message": f"주소 불일치 (입력: {full_address}, 첫결과: {items[0].get('address', {}).get('parcel', '')})"}), 400
                
        except Exception as e:
            print(f"VWorld 주소 검색 오류: {e}")
            return jsonify({"success": False, "message": f"서버 내부 오류: {str(e)}"}), 500
            
        if not pnu:
            return jsonify({"success": False, "message": "주소에서 고유번호(PNU)를 찾을 수 없습니다."}), 400
    else:
        bcode = data.get('bcode')
        san = data.get('san')
        bonbeon = data.get('bonbeon')
        bubeon = data.get('bubeon')
        
        if not bcode or not bonbeon:
            return jsonify({"success": False, "message": "bcode와 bonbeon은 필수입니다."}), 400
        pnu = f"{bcode}{san}{bonbeon.zfill(4)}{bubeon.zfill(4)}"

    try:
        user_area = float(data.get('area', 0))
    except (ValueError, TypeError):
        user_area = 0.0
    
    actual_area = str(user_area) if user_area > 0 else ""
    jimok = "대" if san == '1' else "임"
    zoning_list = []
    
    # 1. VWorld API 실제 연동
    if not VWORLD_KEY:
        return jsonify({"success": False, "message": "서버에 VWorld API Key가 설정되지 않았습니다."}), 500

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7"
    }

    # (1) 토지특성정보 조회
    try:
        url_char = f"http://api.vworld.kr/ned/data/getLandCharacteristics?key={VWORLD_KEY.strip()}&domain=http://127.0.0.1&pnu={pnu}&format=json&numOfRows=50&pageNo=1"
        res_char = requests.get(url_char, timeout=10).json()
        if 'landCharacteristicss' in res_char and 'field' in res_char['landCharacteristicss']:
            fields = res_char['landCharacteristicss']['field']
            if fields:
                latest = sorted(fields, key=lambda x: x.get('stdrYear', '0'))[-1]
                jimok = latest.get('lndcgrCodeNm', jimok)
                real_area = latest.get('lndpclAr', '')
                if real_area and user_area <= 0:
                    actual_area = str(real_area)
    except Exception as e:
        print(f"VWorld 토지특성정보 통신 오류: {e}")

    # (2) 토지이용계획(지역지구) 실데이터 조회
    try:
        url_zoning = f"http://api.vworld.kr/ned/data/getLandUseAttr?key={VWORLD_KEY.strip()}&domain=http://127.0.0.1&pnu={pnu}&format=json&numOfRows=50&pageNo=1"
        res_zoning = requests.get(url_zoning, timeout=10).json()
            
        if 'landUses' in res_zoning:
            if 'field' in res_zoning['landUses']:
                fields = res_zoning['landUses']['field']
                for f in fields:
                    z_name = f.get('prposAreaDstrcCodeNm')
                    if z_name and z_name not in zoning_list:
                        zoning_list.append(z_name)
            else:
                err_msg = res_zoning['landUses'].get('resultMsg', '알 수 없는 VWorld 오류')
                zoning_list.append(f"API 에러: {err_msg}")
    except Exception as e:
        zoning_list.append(f"통신 에러: {str(e)}")
        print(f"VWorld 토지이용계획 통신 오류: {e}")
            
    # 2. 결과 조합 (API가 실패했거나 결과가 없을 경우 대비 Fallback)
    if not zoning_list:
        zoning_list.append("지역지구 데이터 없음")
            
    return jsonify({
        "success": True,
        "pnu": pnu,
        "actualArea": actual_area,
        "jimok": jimok,
        "zoning": zoning_list
    })

def run_analysis(job_id, data):
    try:
        genai.configure(api_key=GEMINI_KEY)
        # ⚠️ genai.list_models()는 최근 구글이 모델 목록 API 응답에 추가한 "thinking" 필드를
        # 구버전(EOL 예정) SDK가 파싱하지 못해 TypeError로 죽는다 (앱 전체 공통 문제).
        # 다른 기능들은 전부 이 호출을 try/except로 감싸 고정 목록으로 폴백하는데
        # run_analysis(law_review)만 방어 코드가 없어서 분석 자체가 통째로 실패하고
        # 있었다. 다른 곳과 동일한 방어 패턴 적용.
        try:
            available_models = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods]
            if not available_models:
                raise ValueError("empty model list")
        except Exception as e:
            print(f"[run_analysis] genai.list_models() failed, using fallback list: {e}")
            available_models = ['models/gemini-2.5-pro', 'models/gemini-2.5-flash', 'models/gemini-1.5-pro', 'models/gemini-1.5-flash', 'models/gemini-2.0-flash']

        flash_model_name = None
        pro_model_name = None
        
        for preferred in ['models/gemini-2.5-flash', 'models/gemini-2.0-flash', 'models/gemini-1.5-flash']:
            if preferred in available_models:
                flash_model_name = preferred
                break
                
        for preferred in ['models/gemini-2.5-pro', 'models/gemini-1.5-pro']:
            if preferred in available_models:
                pro_model_name = preferred
                break
                
        if not flash_model_name: flash_model_name = available_models[0]
        if not pro_model_name: pro_model_name = flash_model_name # Fallback to flash if no pro

        extractor_model = genai.GenerativeModel(flash_model_name)
        
        model = genai.GenerativeModel(pro_model_name)
        
        project_name = data.get('projectName', '이름 없음')
        project_type = data.get('projectType', '복합공사')
        budget = data.get('budget', 0)
        budget_nat = data.get('budgetNational', 0)
        budget_prov = data.get('budgetProvincial', 0)
        budget_mun = data.get('budgetMunicipal', 0)
        total_area = data.get('totalArea', 0)
        public_water_area = data.get('publicWaterArea', 0)
        description = data.get('description', '설명 없음')
        parcels = data.get('parcels', [])

        if not parcels and public_water_area <= 0:
            raise Exception("검증된 편입 필지 또는 공유수면 면적이 없습니다.")

        parcel_str_list = []
        all_zonings = set()
        for p in parcels:
            parcel_str_list.append(f"- {p['address']} (면적: {p['area']}㎡) | 지역지구: {p['zoning']}")
            for z in p['zoning'].split(', '):
                all_zonings.add(z)
                
        parcel_str = "\n".join(parcel_str_list)
        if not parcel_str:
            parcel_str = "지번이 부여된 육상 필지 편입 없음"
            
        zoning_context = ", ".join(list(all_zonings))
        
        public_water_instruction = ""
        if public_water_area > 0:
            public_water_instruction = f"""
        [특수 조건: 해상 및 공유수면 공사]
        본 사업 구역에는 지번이 존재하지 않는 바다/하천 등 **공유수면 면적 {public_water_area}㎡**가 포함되어 있습니다.
        따라서 토지이음 규제 내역에 나타나지 않더라도, 해상 공사 및 공유수면 점용과 관련된 다음 법률 및 절차를 반드시 최우선으로 검토하고 각 단계(Phase)에 포함하십시오:
        - 「공유수면 관리 및 매립에 관한 법률」 (점용·사용 허가, 실시계획 승인)
        - 「해양환경관리법」 (해역이용협의 등)
        - 「해양조사와 해양정보 활용에 관한 법률」 등 해상 인허가 관련 필수 법령
"""
        # 모든 DB 변수 추출
        all_rule_vars = get_all_variables()
        vars_instruction = ", ".join(all_rule_vars)
        
        # --- 에이전트 1: 파라미터 추출기 (Extractor Agent) ---
        extractor_prompt = f"""
        당신은 건설공사 내역 분석 에이전트입니다.
        아래 [사업 개요] 및 [편입필지 지역지구] 정보를 바탕으로 JSON 데이터를 추출하세요.
        응답은 순수 JSON 형식만 반환하세요 (마크다운 백틱 제외).
        
        [사업 개요] 사업명: {project_name}, 공종: {project_type}, 총 사업비: {budget}억, 면적: {total_area}㎡, 주요 내용: {description}
        [지역지구] {zoning_context}
        
        당신은 다음 변수 목록 중에서 사업 개요에 명확히 해당되는(True 또는 숫자값이 존재하는) 변수들만 골라내어 JSON 키-값 쌍으로 만들어야 합니다:
        [변수 목록]: {vars_instruction}
        
        (주의: 정보가 부족하거나 해당되지 않는 변수는 아예 JSON 키를 생성하지 마세요. 불확실한 경우에도 제외하세요. "budget", "total_area", "floors", "excavation_depth"와 같은 숫자 변수는 숫자로 추출하세요. "has_", "is_" 로 시작하는 것은 true/false로 출력하세요.)
        """
        extractor_resp = extractor_model.generate_content(extractor_prompt)
        ext_text = extractor_resp.text.strip().replace('```json', '').replace('```', '').strip()
        
        try:
            extracted_params = json.loads(ext_text)
        except:
            extracted_params = {}
            
        # 파라미터 합성
        kb_params = extracted_params.copy()
        kb_params.update({
            'budget': budget,
            'budget_nat': budget_nat,
            'total_area': total_area,
            'is_public': True
        })
        
        if '임야' in zoning_context or '산지' in zoning_context: kb_params['has_mountain'] = True
        if '농지' in zoning_context or '전' in zoning_context or '답' in zoning_context: kb_params['has_farmland'] = True

        # --- Rule Engine (Knowledge Base 결정론적 매칭) ---
        matched_laws = evaluate_knowledge_base(kb_params)
        scale_permits_str = ""
        if matched_laws:
            law_lines = [f"        - {law['name']} (Phase: {law['phase']}) : {law['desc']} (근거: {law['law_link']})" for law in matched_laws]
            scale_permits_str = "\n".join(law_lines)
            scale_permits_str = f"\n        **[지식 기반(Knowledge Base) 강제 적용 목록]**\n        서버의 룰 엔진이 매칭한 절대 누락되어서는 안 될 필수 목록입니다. 이 항목들은 반드시 최종 보고서 JSON의 적절한 phase 배열에 포함시키세요:\n{scale_permits_str}\n"

        # 국토계획법 등 기본 정보 패치 -> MCP 에이전트 자율 검색으로 업그레이드
        from mcp_agent_sync import get_mcp_context_sync
        mcp_query = f"'{project_type}' 사업에 적용되는 '{zoning_context}' 관련 법령과 필수 인허가 절차를 최대한 상세히 찾아줘."
        if public_water_area > 0:
            mcp_query += " 해상 및 공유수면 공사가 포함되어 있으니 이와 관련된 해양 인허가 법령을 반드시 찾아줘."
            
        mcp_rag_context = get_mcp_context_sync(mcp_query)
        law_context = f"[법제처 API 기반 MCP 에이전트 실시간 검색 결과]\n{mcp_rag_context}"

        prompt = f"""
        당신은 대한민국 시설직 공무원을 돕는 최고 수준의 법규 검토 AI 전문가입니다.
        
        🚨 [절대 누락 금지: 무결점 감사 대비 모드] 🚨
        사용자는 공무원이며, 단 하나의 법적 절차, 인허가, 현장 감독 사항이라도 누락될 경우 심각한 징계나 감사 지적을 받게 됩니다.
        따라서 사업 규모나 내용에 비추어 볼 때 단 1%의 가능성이라도 있는 법정 의무사항이나 인허가는 **절대로 생략하지 말고 모두 빠짐없이 도출**하십시오. 당신은 공무원의 완벽한 업무 처리를 보장하는 최후의 안전망입니다.
        
        아래 [사업 개요]와 토지대장에서 실시간으로 조회한 [편입필지 지역지구] 정보를 바탕으로 분석하세요.

        [사업 개요]
        - 사업명: {project_name}
        - 주요 사업 분류 (공종): {project_type}
        - 총 사업비: {budget}억 원
        - 검증된 총 사업 면적: {total_area}㎡
        - 주요 사업 내용: {description}
        
        [편입 필지 및 지역지구 현황] (매우 중요)
        {parcel_str}
        {public_water_instruction}
        
        ※ 핵심 검토 요건: 이 사업은 [{zoning_context}] 구역을 포함하고 있습니다. 이 용도지역에 따른 행위 제한 및 필수 인허가를 반드시 찾아내어 적으세요.
        {scale_permits_str}
        [법제처 제공 현행법 컨텍스트]
        {law_context}

        **[법령 조항 번호 명시 (매우 중요)]**
        1. 당신이 도출한 필수 법적 절차에 대해, **정확한 조항 번호(예: 제8조, 제10조 제1항 등)를 반드시 기재**하십시오.
        2. 제공된 텍스트(컨텍스트)에 있는 [법제처 API 기반 MCP 에이전트 실시간 검색 결과]의 법률 원문을 최우선으로 적극 참고하십시오. 만약 조문이 충분하지 않다면, **구글 검색 도구(Google Search Retrieval)를 함께 활용**하여 최신 현행 법령을 찾아내십시오.
        3. 🚨 **[현행법령 확인 의무]**: 검색 시 반드시 연혁법령이나 폐지된 법률이 아닌 **"현행법령(현재 시행 중인 법률)"**인지 확인해야 합니다. 만약 옛날 블로그 글이나 구법(폐지된 법)의 조항이라면 절대 사용하지 말고, 현행 법제처 기준으로 재확인하십시오.
        4. 조항 번호를 기재할 때는 절대 "컨텍스트에 없어서~"와 같은 변명이나 사과문, 해명글을 적지 마십시오. 전문적인 보고서 문체만을 유지하세요.

        **[전문가 수준의 심층 연관 분석 프레임워크 (Deep Reasoning Framework)]**
        당신은 단순 정보 검색기가 아니라 최상급 건설 행정 전문가입니다. 아래 3단계 사고 프레임워크를 반드시 거쳐 결과를 도출하십시오.
        
        [1단계: 사업 본질 분해 (Project Deconstruction)]
        제시된 사업 개요(예산, 면적, 공사내용)를 분석하여 해당 사업이 어떤 특성(예: 하천 공사, 산지 전용, 시특법상 1·2·3종 시설물 등)을 갖는지 스스로 규정하십시오.
        
        [2단계: 법적 트리거 탐색 (Legal Trigger Search)]
        분해된 사업 특성을 바탕으로, 당신의 내부 지식(Pre-trained Knowledge)을 총동원하여 '이 조건일 때 발동하는(Trigger) 법적 의무'가 무엇인지 샅샅이 탐색하십시오. 
        특히 「건설기술 진흥법」, 「산업안전보건법」, 「환경영향평가법」 등 대한민국 건설공사 핵심 법령에 따른 안전/환경/품질 관리 의무를 절대 누락하지 마십시오.
        
        [3단계: 생애주기별 역산 (Lifecycle Reverse-Engineering)]
        시공 단계의 의무만 찾지 마십시오. 도출된 의무를 이행하기 위해 기획이나 설계 단계에서 미리 준비해야 하는 행정 절차(예: 설계단계 건설사업관리, 타당성조사, 각종 영향평가, 심의 등)를 역산하여 도출하고 적절한 phase에 배치하십시오.

        **[추가 지시사항: 공종에 따른 엄격한 필터링 및 추측 금지]**
        이 사업은 '{project_type}'입니다.
        - [절대 금지 사항] '토목공사이더라도 건축물이 포함될 경우를 대비하여'와 같은 추측성 판단을 절대 하지 마십시오.
        - 입력된 공종이 '토목공사'라면 사업 현장에 100% 토목 시설물만 존재한다고 확정 지으십시오. 건축법, 건축서비스산업진흥법 등 건축 관련 법령은 단 1%도 도출해서는 안 됩니다.
        - 조경공사 역시 건축 관련 규제는 철저히 배제하고 산지/공원 위주로 도출하십시오.
        - 복합공사라면 모든 가능성을 열어두고 종합적으로 검토하세요.

        [요청 사항]
        보고서에 쓸 수 있도록 전문적인 용어로 답변하되, 응답은 반드시 아래 JSON 형식(마크다운 백틱 없이 순수 JSON만)으로 반환하세요.
        URL 링크 엉킴을 방지하기 위해, 관련 법령은 절대로 HTML 태그를 쓰지 말고 "law_name"(정확한 법령명 띄어쓰기 준수)과 "article"(조항 번호)로 정확히 분리하여 작성해 주세요.
        - 타당성 조사, 기본계획, 투자심사 등 구상/기획 단계에 해당하는 항목은 'planning' 배열에 작성하세요.
        - 건축허가, 도로점용허가 등 인허가 항목은 'permits' 배열에 별도로 분리하여 작성하세요.
        - 대금 지급(선금, 기성금, 준공금 등)과 관련된 절차가 여러 단계에 걸쳐 중복될 경우, 하나로 통합하여 '공사 대금 청구 및 지급' 등의 단일 항목으로 병합하세요.
        {{
            "permits": [
                {{
                    "name": "건축허가",
                    "law_name": "건축법",
                    "article": "제11조",
                    "reason": "해당 사업은 계획관리지역 내 새로운 건축물을 축조하는 사업이므로 건축허가가 필요함."
                }}
            ],
            "phases": {{
                "planning": [
                    {{"task": "타당성 조사 및 투자심사", "law_name": "지방재정법", "article": "제37조", "desc": "신규 투자사업에 대한 예산 편성 전 타당성 조사 및 투자심사 의뢰"}}
                ],
                "design": [
                    {{"task": "설계안전성검토 의뢰", "law_name": "건설기술 진흥법", "article": "제62조", "desc": "가설구조물 및 굴착 공사에 따른 설계안전성 사전 검토 (대상 여부 확인 필요)"}}
                ],
                "construction": [
                    {{"task": "안전관리계획서 제출", "law_name": "건설기술 진흥법", "article": "제62조", "desc": "착공 전 인허가 기관에 안전관리계획서 제출 및 승인"}}
                ],
                "completion": [
                    {{"task": "준공검사 신청", "law_name": "건설기술 진흥법", "article": "제39조", "desc": "공사 완료 후 발주청에 준공검사원 제출"}}
                ],
                "maintenance": [
                    {{"task": "하자보수 점검", "law_name": "건설산업기본법", "article": "제28조", "desc": "하자담보책임기간 내 정기 점검 실시"}}
                ]
            }}
        }}
        """
        
        try:
            response = generate_with_search_grounding(prompt, pro_model_name)
        except Exception as e:
            print(f"Search grounding failed, falling back to standard: {e}")
            response = model.generate_content(prompt)
            
        text_resp = response.text.strip()
        
        if text_resp.startswith("```json"):
            text_resp = text_resp[7:]
        if text_resp.startswith("```"):
            text_resp = text_resp[3:]
        if text_resp.endswith("```"):
            text_resp = text_resp[:-3]
            
        result = json.loads(text_resp.strip())
        
        # 100% Guaranteed Link Injection
        base_dir = os.path.dirname(os.path.abspath(__file__))
        law_urls_path = os.path.join(base_dir, 'data', 'law_urls.json')
        law_urls_dict = {}
        if os.path.exists(law_urls_path):
            with open(law_urls_path, 'r', encoding='utf-8') as f:
                law_urls_dict = json.load(f)
                
        import urllib.parse
        import requests
        import xml.etree.ElementTree as ET
        import concurrent.futures

        def resolve_url(item):
            if 'law_name' not in item or not item['law_name']:
                return
            law_name = item['law_name'].strip()
            
            if law_name in law_urls_dict:
                item['law_url'] = law_urls_dict[law_name]
                return
                
            try:
                res_law = requests.get(f"https://www.law.go.kr/DRF/lawSearch.do?OC={LAW_KEY}&target=law&type=XML&query={urllib.parse.quote(law_name)}", timeout=3)
                root_law = ET.fromstring(res_law.text)
                if int(root_law.findtext('totalCnt', '0')) > 0:
                    item['law_url'] = f"https://www.law.go.kr/법령/{urllib.parse.quote(law_name)}"
                    return
            except: pass
            
            try:
                res_adm = requests.get(f"https://www.law.go.kr/DRF/lawSearch.do?OC={LAW_KEY}&target=admrul&type=XML&query={urllib.parse.quote(law_name)}", timeout=3)
                root_adm = ET.fromstring(res_adm.text)
                if int(root_adm.findtext('totalCnt', '0')) > 0:
                    item['law_url'] = f"https://www.law.go.kr/행정규칙/{urllib.parse.quote(law_name)}"
                    return
            except: pass
            
            is_admrul = law_name.endswith('지침') or law_name.endswith('기준') or law_name.endswith('고시') or law_name.endswith('규정')
            base = 'https://www.law.go.kr/LSW/admRulSc.do?query=' if is_admrul else 'https://www.law.go.kr/LSW/lsSc.do?query='
            item['law_url'] = base + urllib.parse.quote(law_name)
            
        items_to_resolve = []
        if 'permits' in result:
            items_to_resolve.extend(result['permits'])
        if 'phases' in result:
            for phase_items in result['phases'].values():
                items_to_resolve.extend(phase_items)
                
        with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
            executor.map(resolve_url, items_to_resolve)

        # AI가 구글 검색 그라운딩 없이 자체 지식만으로 답한 항목에서, 존재하지 않거나
        # 이미 폐지된 조항을 현행처럼 인용하는 사고가 실사용 중 발견됨(예: 국토계획법
        # 제71조 - 2006년 삭제 - 를 현행 조항처럼 인용). 로컬에 원문이 있는 법(62개)에
        # 한해서만이라도 조항 실존 여부를 기계적으로 대조해 걸러낸다.
        for item in items_to_resolve:
            _verify_citation(item)

        JOBS[job_id] = {"status": "completed", "result": result}
        
    except json.JSONDecodeError as e:
        JOBS[job_id] = {"status": "error", "message": "AI가 반환한 데이터를 파싱할 수 없습니다."}
    except Exception as e:
        JOBS[job_id] = {"status": "error", "message": str(e)}

@app.route('/api/analyze/start', methods=['POST'])
def analyze_start():
    if not GEMINI_KEY:
        return jsonify({"error": "Google Gemini API 키가 설정되지 않았거나 만료되었습니다. 클라우드타입(Cloudtype) 설정 -> 환경변수에서 'GEMINI_API_KEY'를 추가한 후 재배포해주세요."}), 400

    data = request.json
    job_id = str(uuid.uuid4())
    JOBS[job_id] = {"status": "processing"}
    
    thread = threading.Thread(target=run_analysis, args=(job_id, data))
    thread.start()
    
    return jsonify({"success": True, "job_id": job_id})

@app.route('/api/analyze/status/<job_id>', methods=['GET'])
def analyze_status(job_id):
    if job_id not in JOBS:
        return jsonify({"status": "error", "message": "존재하지 않는 작업입니다."}), 404
        
    job_info = JOBS[job_id]
    
    if job_info["status"] == "completed":
        result = job_info.get("result", {})
        # 메모리 정리를 위해 완료된 작업은 삭제
        del JOBS[job_id]
        return jsonify({"status": "completed", "result": result})
        
    elif job_info["status"] == "error":
        error_msg = job_info.get("message", "알 수 없는 오류")
        del JOBS[job_id]
        return jsonify({"status": "error", "message": error_msg})
        
    return jsonify({"status": "processing"})



import tempfile
import werkzeug.utils


def run_other_review(job_id, text_content, temp_path, filename, file_obj_exists):
    try:
        import os
        import tempfile
        import google.generativeai as genai
        
        genai.configure(api_key=GEMINI_KEY)
        
        uploaded_file = None
        file_text = ""
        
        if temp_path and os.path.exists(temp_path):
            try:
                uploaded_file = upload_file_for_gemini(temp_path, display_name=filename)
            except Exception as e:
                print(f"genai.upload_file failed: {e}")
                uploaded_file = None
                
            # 항상 로컬 파이썬 환경에서도 텍스트를 추출 (RAG 및 정규식 스캔용)
            try:
                import fitz
                doc = fitz.open(temp_path)
                for page in doc:
                    file_text += page.get_text() + "\n"
                doc.close()
            except Exception as e:
                print(f"PyMuPDF error: {e}")
                try:
                    import PyPDF2
                    with open(temp_path, 'rb') as f:
                        reader = PyPDF2.PdfReader(f)
                        for page in reader.pages:
                            text = page.extract_text()
                            if text: file_text += text + "\n"
                except Exception as e2:
                    print(f"PyPDF2 error: {e2}")
                    try:
                        with open(temp_path, 'r', encoding='utf-8') as f:
                            file_text = f.read()
                    except:
                        pass
            
            # genai.upload_file이 실패했을 때만 프롬프트에 직접 텍스트 첨부
            if not uploaded_file and file_text:
                text_content += f"\n\n[첨부 문서 내용]\n{file_text[:30000]}"

            try:
                os.remove(temp_path)
            except Exception as e:
                print(f"temp file cleanup failed: {e}")
            
        try:
            available_models = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods and 'vision' not in m.name.lower()]
            models_to_try = sorted(available_models, key=lambda x: (0 if '1.5-pro' in x else 1 if '2.5-pro' in x else 2 if 'pro' in x else 3 if '1.5-flash' in x else 4))
        except:
            models_to_try = ['models/gemini-2.5-pro', 'models/gemini-1.5-pro', 'models/gemini-2.5-flash', 'models/gemini-2.0-flash']
            
        # RAG 검색 시에는 파일 내 텍스트까지 합쳐서 분석!
        full_query_for_rag = text_content + "\n" + file_text
            
        from mcp_agent_sync import get_mcp_context_sync
        
        # Call the new MCP Agent to search both laws and precedents
        mcp_rag_context = get_mcp_context_sync(full_query_for_rag)
        
        moleg_context = f"[법제처 API 실시간 RAG 검색 결과 (MCP Agent)]\n{mcp_rag_context}"
        precedent_context = "" # Merged into moleg_context
        local_law_context = fetch_local_law_data(full_query_for_rag, moleg_context)
        
        raw_text = text_content.strip()
        has_file = bool(file_obj_exists or file_text or uploaded_file)
        dummy_phrases = ["이네용", "이거", "이거 분석해줘", "분석해줘", "검토해줘"]
        has_text = bool(raw_text and raw_text not in dummy_phrases and len(raw_text) > 5)

        if has_file and has_text:
            mode_instruction = "🚨 **[분석 모드: 질의 + 첨부문서]** 사용자가 특정 질문/지시사항과 함께 문서를 첨부했습니다. 첨부된 문서를 꼼꼼히 분석하되, 반드시 **사용자가 입력한 [요청 내용]을 최우선으로 반영하여 실행/답변**하십시오. 필요시 아래 4단계 구조를 적절히 활용하여 보충하십시오."
        elif has_file and not has_text:
            mode_instruction = "🚨 **[분석 모드: 문서 단독 분석]** 사용자가 특별한 질문 없이 문서만 첨부했습니다. 절대로 '질문이 부족하다'고 핑계대지 마시고, 첨부된 문서 자체를 심층 분석하여 스스로 핵심 쟁점을 도출해낸 뒤 **무조건 아래 4단계 구조의 종합 법리 검토 보고서를 작성**하십시오."
        else:
            mode_instruction = "🚨 **[분석 모드: 일반 질의응답]** 사용자가 첨부문서 없이 일반적인 질문이나 지시를 내렸습니다. [요청 내용]을 분석하여 정확한 법리적 답변 및 지시된 포맷으로 제공하십시오. 가급적 아래 4단계 구조를 따르되, 단순한 질문이나 특정 지시가 있다면 그 지시에 맞게 유연하게 답변하십시오."
        
        prompt = f"""당신은 대한민국 공무원들의 행정, 감사, 예산 업무를 지원하는 '다중 에이전트(법무/감사/재무 전문가)'입니다.
공무원이 다음 상황에 대한 검토를 요청했습니다.

[사용자 질문]
{text_content}

{moleg_context}

{precedent_context}

[로컬 법령 데이터베이스 (조문 본문 및 별표/서식)]
{local_law_context}

[특별 지시사항]
1. [법제처 API 실시간 RAG 검색 결과]와 [로컬 법령 데이터베이스]를 최우선으로 인용하십시오. 특히 별표(Attached Tables)에 대한 질문은 로컬 데이터베이스의 내용을 바탕으로 상세히 설명하십시오. RAG에 포함된 하이퍼링크를 그대로 사용하세요.
   - 만약 [자치법규(조례) 검색 결과]가 함께 제공되었다면, 이는 사용자가 언급한 지방자치단체의 실제 조례/규칙 검색 결과입니다. 사안과 관련된 조례가 있다면 반드시 그 조례명과 링크를 본문(2. 핵심 쟁점, 3. 관련 법령, 4. 행동 지침 어디든 관련되는 곳)에 인용하십시오. "OO 조례를 확인해야 합니다"라고 막연히 안내만 하지 말고, 검색 결과에 실제로 존재하는 조례가 있다면 그 이름과 링크를 직접 제시하십시오. 검색 결과가 없거나 제공되지 않았다면 그때만 "자치법규정보시스템(elis.go.kr) 또는 해당 지자체에 직접 확인이 필요합니다"라고 명시하십시오.
2. 법령/조례 조문, 판례, 해석례를 인용할 때는 아래의 규칙을 **보고서 전체 어디에서든(관련 법령 섹션뿐 아니라 상황 요약·핵심 쟁점·행동 지침·결론까지 모두) 예외 없이** 완벽하게 준수하십시오. 이미 앞에서 링크로 인용한 법령이라도 뒤에서 다시 언급할 때 맨 텍스트로 쓰지 말고 매번 동일하게 링크로 쓰십시오.
  - ⚖️ 법령 조문 링크: 법률의 성격에 따라 URL 형식을 엄격히 구분하여 반드시 **제X조**까지 구체적으로 연결되도록 작성하십시오. (띄어쓰기는 그대로 유지합니다)
    - **법률, 시행령, 시행규칙**인 경우 (예: ~법, ~령, ~규칙): `[법령명 제X조](https://www.law.go.kr/법령/법령명/제X조)`
    - **고시, 훈령, 예규, 지침, 기준**인 경우 (예: ~고시, ~기준, ~지침): `[행정규칙명 제X조](https://www.law.go.kr/행정규칙/행정규칙명/제X조)`
    - **지방자치단체 조례, 규칙**인 경우 (예: ~조례, ~조례 시행규칙): `[조례명](https://www.law.go.kr/자치법규/조례명)` (조례는 [자치법규(조례) 검색 결과]에 제공된 링크를 그대로 사용하고, 직접 조항 번호까지는 지어내지 마십시오.)
3. 🔎 **[판례 인용] 제공된 판례 데이터베이스 활용**:
  - 만약 사용자가 첨부한 문서에 특정한 사건번호(판례)가 인용되어 있다면, [법제처 실시간 판례 검색 결과] 중 **[문서 내 인용된 사건번호 추적 결과]**를 최우선으로 확인하고 답변에 분석/반영하십시오. (가짜 판례로 판명된 경우 그 사실을 지적하십시오.)
  - 제공된 원문 판례를 100% 신뢰하여 답변에 인용하십시오.
  - 인용 시 반드시 제공된 하이퍼링크 형식 `[판례명(사건번호)](링크)`을 그대로 유지하여 클릭할 수 있도록 만드십시오.
  - 🚨 **[가짜 판례 창작 절대 금지]**: 제공된 데이터베이스에 관련된 판례가 없다면, 절대 임의로 사건번호를 지어내거나(할루시네이션) 인터넷 검색을 시도하지 마십시오. 판례가 제공되지 않은 경우 "현재 쟁점과 관련된 대법원 판례 데이터가 제공되지 않았습니다."라고만 출력하십시오.
4. 응답 구조 및 모드:
  {mode_instruction}
5. 🚨 **[할루시네이션(환각) 원천 차단]**: 만약 사용자가 질의한 특정 법률의 원문이 [로컬 법령 데이터베이스]에 제공되지 않았다면, **절대로 조항 번호(예: 제X조)나 구체적 내용을 스스로 창작하거나 유추해서 적지 마십시오!** 이 경우 일반적인 법리와 절차만 설명하고, 반드시 "해당 법률의 원문 데이터가 로컬 DB에 없어 정확한 조항 번호는 법제처(law.go.kr)를 직접 참조하시기 바랍니다."라고 명시하십시오.

### 1. 상황 요약 (Situation Summary)
- 
### 2. 핵심 쟁점 (Key Legal Issues)
- 
### 3. 관련 법령 및 핵심 판례 (Applicable Laws & Key Precedents)
- 법령은 [법제처 API 실시간 RAG 검색 결과] 및 [로컬 법령 데이터베이스]를 바탕으로 상세 설명. 
- 판례는 [법제처 실시간 판례 검색 결과]에 제시된 판례만 인용하여 요점을 상세히 설명하십시오. (제공된 결과가 없으면 "관련 판례 없음"만 명시할 것.)
### 4. 공무원 행동 지침 및 결론 (Action Plan)
- 

[요청 내용]
{text_content}"""
        
        contents_payload = [prompt]
        if uploaded_file:
            contents_payload.append(gemini_file_part(uploaded_file))

        response = None
        last_err = None
        for m in models_to_try:
            try:
                try:
                    model = genai.GenerativeModel(model_name=m)
                    response = model.generate_content(contents_payload)
                    break
                except Exception as tool_e:
                    print(f"Tool {m} fallback: {tool_e}")
                    model = genai.GenerativeModel(model_name=m)
                    response = model.generate_content(contents_payload)
                    break
            except Exception as e:
                last_err = e
                print(f"Model {m} failed: {e}")
                continue
                
        if not response:
            raise Exception(f"모든 AI 모델이 요청 한도 초과 또는 오류로 실패했습니다. 마지막 오류: {last_err}")
            
        file_name = uploaded_file.name if uploaded_file else ""
        JOBS[job_id] = {
            "status": "completed",
            "result": response.text,
            "file_name": file_name,
            "initial_context": full_query_for_rag
        }
    except Exception as e:
        print(f"run_other_review error: {e}")
        JOBS[job_id] = {"status": "error", "message": str(e)}

@app.route('/api/analyze/other_review', methods=['POST'])
def api_other_review():
    try:
        import tempfile
        import werkzeug.utils
        import uuid
        import os
        import threading
        
        if request.content_type and 'multipart/form-data' in request.content_type:
            text_content = request.form.get('text', '')
            file_obj = request.files.get('file')
        else:
            data = request.json or {}
            text_content = data.get('text', '')
            file_obj = None
            
        if not text_content and not file_obj:
            return jsonify({"success": False, "message": "검토할 내용이나 파일이 제공되지 않았습니다."}), 400
            
        temp_path = None
        filename = None
        file_obj_exists = False
        
        if file_obj and file_obj.filename:
            file_obj_exists = True
            filename = werkzeug.utils.secure_filename(file_obj.filename)
            temp_dir = tempfile.gettempdir()
            temp_path = os.path.join(temp_dir, f"{uuid.uuid4()}_{filename}")
            file_obj.save(temp_path)
            
        job_id = str(uuid.uuid4())
        JOBS[job_id] = {"status": "processing"}
        
        thread = threading.Thread(target=run_other_review, args=(job_id, text_content, temp_path, filename, file_obj_exists))
        thread.start()
        
        return jsonify({"success": True, "job_id": job_id})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/analyze/other_review_status/<job_id>', methods=['GET'])
def other_review_status(job_id):
    if job_id not in JOBS:
        return jsonify({"status": "error", "message": "존재하지 않는 작업입니다."}), 404
    
    job_info = JOBS[job_id]
    
    if job_info["status"] == "completed":
        result = job_info.get("result", "")
        file_name = job_info.get("file_name", "")
        initial_context = job_info.get("initial_context", "")
        del JOBS[job_id]
        return jsonify({
            "status": "completed", 
            "result": result,
            "file_name": file_name,
            "initial_context": initial_context
        })
        
    elif job_info["status"] == "error":
        error_msg = job_info.get("message", "알 수 없는 오류")
        del JOBS[job_id]
        return jsonify({"status": "error", "message": error_msg})
        
    else:
        return jsonify({"status": "processing"})

@app.route('/api/chat/other_review', methods=['POST'])
def api_chat_other_review():
    try:
        data = request.json
        chat_history = data.get('chat_history', [])
        new_message = data.get('new_message', '')
        file_name = data.get('file_name', '')
        initial_context = data.get('initial_context', '')
        
        if not new_message:
            return jsonify({"success": False, "message": "질문이 제공되지 않았습니다."}), 400
            
        genai.configure(api_key=GEMINI_KEY)
        
        contents_payload = []
        first_user_parts = [initial_context]
        if file_name:
            try:
                file_obj = genai.get_file(file_name)
                first_user_parts.append(file_obj)
            except Exception as e:
                print(f"File retrieval failed: {e}")
                
        contents_payload.append({"role": "user", "parts": first_user_parts})
        
        for msg in chat_history:
            contents_payload.append({
                "role": msg["role"],
                "parts": [msg["text"]]
            })
            
        contents_payload.append({
            "role": "user",
            "parts": [new_message]
        })
        
        try:
            available_models = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods and 'vision' not in m.name.lower()]
            models_to_try = sorted(available_models, key=lambda x: (0 if '1.5-pro' in x else 1 if '2.5-pro' in x else 2 if 'pro' in x else 3 if '1.5-flash' in x else 4))
        except:
            models_to_try = ['models/gemini-2.5-pro', 'models/gemini-1.5-pro', 'models/gemini-2.5-flash', 'models/gemini-2.0-flash']
            
        response = None
        last_err = None
        for m in models_to_try:
            try:
                model = genai.GenerativeModel(model_name=m)
                response = model.generate_content(contents_payload)
                break
            except Exception as e:
                last_err = e
                continue
                
        if not response:
            raise Exception(f"채팅 응답 생성 실패: {last_err}")
            
        return jsonify({"success": True, "result": response.text})
        
    except Exception as e:
        print(f"Chat Review API Error: {e}")
        return jsonify({"success": False, "message": f"서버 오류: {str(e)}"})


@app.route('/api/chat/duty_list', methods=['POST'])
def api_chat_duty_list():
    """
    '법정 의무 마스터' 페이지의 법령 상담 챗봇. 프론트(duty_list.html)가 예전부터 이 경로를
    호출하고 있었는데 라우트 자체가 없어서 항상 404였던 것 + 프론트의 변수 스코프 버그가 겹쳐서
    채팅 기능이 완전히 죽어있었다. 둘 다 고침 (스코프는 duty_list.html에서).
    """
    try:
        data = request.json or {}
        chat_history = data.get('chat_history', [])
        new_message = data.get('new_message', '')
        law_name = data.get('law_name', '')
        lsi_seq = data.get('lsi_seq', '')

        if not new_message:
            return jsonify({"success": False, "message": "질문이 제공되지 않았습니다."}), 400

        if not GEMINI_KEY:
            return jsonify({"success": False, "message": "Gemini API 키가 설정되지 않았습니다."}), 500

        # 체크리스트 분석 때 이미 캐시된 법령 원문이 있으면 재사용, 없으면 새로 조회
        global LAW_TEXT_CACHE
        law_text = LAW_TEXT_CACHE.get(lsi_seq, '') if lsi_seq else ''
        if not law_text and lsi_seq:
            try:
                doc_res = requests.get(f"https://www.law.go.kr/DRF/lawService.do?OC={LAW_KEY}&target=law&type=XML&MST={lsi_seq}", timeout=10)
                doc_root = ET.fromstring(doc_res.text)
                articles = doc_root.findall('.//조문단위')
                fetched_text = ""
                for art in articles:
                    art_title = art.findtext('조문내용') or ""
                    fetched_text += art_title + "\n"
                    for hang in art.findall('.//항내용'):
                        if hang.text:
                            fetched_text += hang.text + "\n"
                    for ho in art.findall('.//호내용'):
                        if ho.text:
                            fetched_text += ho.text + "\n"
                law_text = fetched_text
                LAW_TEXT_CACHE[lsi_seq] = law_text
            except Exception as e:
                print(f"chat_duty_list law fetch failed: {e}")

        genai.configure(api_key=GEMINI_KEY)

        if law_text:
            system_context = f"""당신은 '{law_name}' 법령 전문 상담 AI입니다. 아래 [법령 원문]만을 근거로 사용자의 질문에 답변하세요.
원문에 없는 내용은 절대 추측하거나 지어내지 말고 "법령 원문에서 해당 내용을 확인할 수 없습니다"라고 답하세요.
답변 시 관련 조 번호를 반드시 함께 밝히세요.

[법령 원문]
{law_text[:30000]}"""
        else:
            system_context = f"""당신은 '{law_name}' 법령 전문 상담 AI입니다. 법령 원문을 불러오지 못했으니, 확실하지 않은 조문 번호나 세부 수치는
절대 단정하지 말고 "정확한 조항은 법제처 원문을 직접 확인하시기 바랍니다"라고 안내하며 답변하세요."""

        contents_payload = [{"role": "user", "parts": [system_context]}]
        for msg in chat_history:
            contents_payload.append({"role": msg["role"], "parts": [msg["text"]]})
        contents_payload.append({"role": "user", "parts": [new_message]})

        try:
            available_models = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods and 'vision' not in m.name.lower()]
            models_to_try = sorted(available_models, key=lambda x: (0 if '2.5-flash' in x else 1 if '2.5-pro' in x else 2 if 'pro' in x else 3))
        except:
            models_to_try = ['models/gemini-2.5-flash', 'models/gemini-2.5-pro', 'models/gemini-1.5-flash']

        response = None
        last_err = None
        for m in models_to_try:
            try:
                model = genai.GenerativeModel(model_name=m)
                response = model.generate_content(contents_payload)
                break
            except Exception as e:
                last_err = e
                continue

        if not response:
            raise Exception(f"채팅 응답 생성 실패: {last_err}")

        return jsonify({"success": True, "result": response.text})

    except Exception as e:
        print(f"Chat Duty List API Error: {e}")
        return jsonify({"success": False, "message": f"서버 오류: {str(e)}"})

@app.route('/api/chat/report', methods=['POST'])
def api_chat_report():
    try:
        data = request.json
        chat_history = data.get('chat_history', [])
        new_message = data.get('new_message', '')
        report_data = data.get('report_data', {})
        
        if not new_message:
            return jsonify({"success": False, "message": "질문이 제공되지 않았습니다."}), 400
            
        genai.configure(api_key=GEMINI_KEY)
        
        initial_context = f"""
        당신은 건설사업의 [AI 법규 검토 종합 보고서]에 대해 설명해주는 AI 챗봇입니다.
        아래는 당신이 조금 전 분석하여 생성한 보고서의 원본 데이터(JSON)입니다.
        사용자가 이 보고서의 내용에 대해 질문하면, 아래 데이터를 바탕으로 친절하고 전문적으로 답변해 주세요.
        없는 내용을 꾸며내지 마시고, 데이터에 있는 법적 근거와 현장 상황을 연결하여 설명해 주십시오.

        [보고서 데이터]:
        {json.dumps(report_data, ensure_ascii=False, indent=2)}
        """
        
        contents_payload = []
        contents_payload.append({"role": "user", "parts": [initial_context]})
        
        for msg in chat_history:
            contents_payload.append({
                "role": msg["role"],
                "parts": [msg["text"]]
            })
            
        contents_payload.append({
            "role": "user",
            "parts": [new_message]
        })
        
        try:
            available_models = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods]
            if not available_models:
                raise ValueError("empty model list")
        except Exception as e:
            print(f"[api_chat_report] genai.list_models() failed, using fallback list: {e}")
            available_models = ['models/gemini-2.5-flash', 'models/gemini-2.5-pro', 'models/gemini-1.5-flash', 'models/gemini-2.0-flash']

        model_name = None
        for preferred in ['models/gemini-2.5-flash', 'models/gemini-2.0-flash', 'models/gemini-1.5-flash']:
            if preferred in available_models:
                model_name = preferred
                break

        # If no preferred model is found, just pick the very first available one
        if not model_name and available_models:
            model_name = available_models[0]
        elif not model_name:
            model_name = 'models/gemini-2.5-flash'  # Final fallback if list is empty

        model = genai.GenerativeModel(model_name)
        response = model.generate_content(contents_payload)

        return jsonify({
            "success": True,
            "result": response.text
        })

    except Exception as e:
        print(f"Report Chat API Error: {e}")
        return jsonify({"success": False, "message": f"서버 오류: {str(e)}"})

# saved_files의 category 값 -> 사람이 읽을 라벨 (프롬프트/로그용)
DESIGN_DOC_LABELS = {
    'file_report': '설계보고서',
    'file_estimate': '설계내역서(공사비)',
    'file_quantity': '물량산출서(수량산출서)',
    'file_drawing': '설계도면',
}


def run_design_review(job_id, project_name, project_domain, review_modes, additional_notes, saved_files):
    try:
        import os
        import google.generativeai as genai
        genai.configure(api_key=GEMINI_KEY)

        # 카테고리(문서 종류)별로 텍스트를 따로 모아둔다 - 뒤에서 "도면 vs 물량산출서
        # vs 내역서" 3자 대조를 프롬프트에 명시적으로 지시하려면, 어느 텍스트가
        # 어느 문서에서 나왔는지 구분이 돼 있어야 하기 때문(요청: "도면을 보고 물량을
        # 산출했을 때 내가 주는 물량산출내역서와 같은지 다른지를 알고싶어").
        texts_by_category = {cat: "" for cat in DESIGN_DOC_LABELS}
        filenames_by_category = {cat: [] for cat in DESIGN_DOC_LABELS}
        extracted_text_combined = ""
        uploaded_genai_files = []

        for file_key, file_info in saved_files.items():
            path = file_info['path']
            fname = file_info['name']
            category = file_info.get('category', 'file_report')
            if os.path.exists(path):
                try:
                    gfile = upload_file_for_gemini(path, display_name=fname)
                    uploaded_genai_files.append(gfile)
                except Exception as e:
                    print(f"genai upload failed for {fname}: {e}")

                text = extract_text_from_file(path)
                texts_by_category[category] = texts_by_category.get(category, "") + text
                filenames_by_category.setdefault(category, []).append(fname)
                extracted_text_combined += text

                try:
                    os.remove(path)
                except:
                    pass

        kcsc_context = ""
        if kcsc_engine:
            kcsc_context = kcsc_engine.build_kcsc_context_for_llm(project_name, project_domain, review_modes, extracted_text_combined)

        guideline_context = ""
        try:
            from guideline_store import build_guideline_context_for_llm
            guideline_context = build_guideline_context_for_llm()
        except Exception as e:
            print(f"[Design Review] guideline_store 로드 실패: {e}")

        try:
            available_models = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods and 'vision' not in m.name.lower()]
            models_to_try = sorted(available_models, key=lambda x: (0 if '1.5-pro' in x else 1 if '2.5-pro' in x else 2 if 'pro' in x else 3 if '1.5-flash' in x else 4))
        except:
            models_to_try = ['models/gemini-1.5-pro', 'models/gemini-2.5-pro', 'models/gemini-1.5-flash']

        has_drawing = len(filenames_by_category['file_drawing']) > 0
        has_quantity_doc = len(filenames_by_category['file_quantity']) > 0
        has_estimate_doc = len(filenames_by_category['file_estimate']) > 0

        # 도면에서 AI가 직접 읽어낸 개략 물량을, 사용자가 제공한 물량산출서/내역서
        # 수량과 항목별로 대조하는 표를 만들도록 명시적으로 지시. 도면이나 비교
        # 대상 문서가 없으면 이 표 자체를 건너뛰라고 못박아서, 자료도 없이 표를
        # 지어내는 것을 막는다.
        if has_drawing and (has_quantity_doc or has_estimate_doc):
            compare_targets = []
            if has_quantity_doc:
                compare_targets.append("물량산출서")
            if has_estimate_doc:
                compare_targets.append("내역서")
            quantity_check_instruction = f"""
        [물량 대조 검증 - 반드시 별도 표로 작성]
        - 첨부된 '설계도면'의 치수·개수·면적·연장 등을 직접 판독하여, 확인 가능한 주요 공종별 개략 물량을 산출하십시오
          (예: 굴착/터파기 물량 m³, 콘크리트 물량 m³, 철근 물량 ton, 포장 면적 m², 배관 연장 m 등 도면에서 실제로 치수를 딸 수 있는 항목).
        - 이렇게 도면에서 직접 산출한 값을, 첨부된 {' 및 '.join(compare_targets)}에 기재된 수량과 항목별로 대조하십시오.
        - 반드시 아래 표 형식으로 "📐 물량 대조 결과"라는 별도 섹션에 제시하십시오:

          | 공종/항목 | 도면 기준 AI 산출물량 (산출근거) | 물량산출서 기재값 | 내역서 기재값 | 대조 결과 |
          |---|---|---|---|---|
          | (예: 흙막이 벽체 면적) | 000 m² (OO도면, 길이 00m x 높이 00m 기준) | 000 m² | 000 m² | 🟢 일치 / 🟡 차이 N% / 🔴 확인불가 |

        - "산출근거" 칸에는 반드시 어느 도면의 어떤 치수를 근거로 계산했는지 구체적으로 밝히십시오. 근거를 특정할 수 없으면 그 항목은 "🔴 확인불가"로 표기하고 임의로 숫자를 만들어내지 마십시오.
        - 도면 판독으로 산출한 물량은 AI의 육안 판독에 의한 "개략치"이며, 실제 CAD 수치/현장 실측과 다를 수 있다는 점을 이 표 바로 위에 명시하십시오.
        - 차이가 5% 이상 나는 항목은 🟡 또는 🔴 등급 판정 목록에도 반드시 함께 반영하십시오.
        """
        else:
            missing = []
            if not has_drawing:
                missing.append("설계도면")
            if not (has_quantity_doc or has_estimate_doc):
                missing.append("물량산출서/내역서")
            quantity_check_instruction = f"""
        [물량 대조 검증]
        - {' 및 '.join(missing)}이(가) 첨부되지 않아 물량 대조가 불가능합니다. "📐 물량 대조 결과" 섹션에는 어떤 자료가 부족해서 대조를 못 했는지만 명시하고, 표는 만들지 마십시오(자료 없이 물량을 추정해서 표를 지어내지 마십시오).
        """

        system_instruction = f"""
        당신은 대한민국 최고 권위의 국가기술자격 기술사(토목/건축/안전)이자 건설공사 설계도서 심사 위원장입니다.
        제공된 설계 성과물(설계보고서/설계내역서/물량산출서/설계도면)과 KCSC 국가건설기준(KDS/KCS) 데이터, 그리고 업로드된 부처별 실무 지침을 바탕으로 정밀 교차 검증을 수행하세요.

        [검증 및 작성 원칙]
        1. 반드시 아래의 4단계 등급화 판정 체계로 명확히 구분하여 작성하세요:
           - 🔴 [법규/지침 위반 및 필수 누락 사항] (감사 지적 1순위, 법정 경비 요율 미달, 시방서/부처 지침 규격 위반, 물량 중대 불일치 등)
           - 🟡 [도면 ↔ 내역서 ↔ 물량산출서 ↔ 보고서 간 불일치 의심 항목] (수량 상이, 공법 표기 불일치, 누락 공종 등)
           - 🟢 [적정 및 우수 반영 사항] (KCSC 기준, 부처 지침 및 안전 지침 준수 항목)
           - 💡 [KCSC 표준시방서 및 부처 지침 기반 개선 권고사항] (품질 향상 및 시공성 개선을 위한 제언)
        2. 지적 사항에는 반드시 구체적인 법적 근거(고시명, 법조항) 또는 국가건설기준 코드(예: KDS 21 30 00, KCS 14 20 00) 또는 업로드된 지침 문서명을 명시하세요.
        3. {quantity_check_instruction}
        4. 실무 전문가답게 격식 있고 구체적인 한국어로 작성하세요.
        """

        doc_summary = "\n".join(
            f"  - {DESIGN_DOC_LABELS[cat]}: {', '.join(filenames_by_category[cat]) if filenames_by_category[cat] else '(첨부 없음)'}"
            for cat in DESIGN_DOC_LABELS
        )

        prompt = f"""
        [공사 기본 정보]
        - 공사명: {project_name}
        - 공종 분야: {project_domain}
        - 집중 검토 모드: {', '.join(review_modes)}
        - 추가 질의 사항: {additional_notes or '없음'}

        [첨부된 문서 종류]
        {doc_summary}

        {kcsc_context}

        {guideline_context}

        [추출된 문서 요약/텍스트 일부 - 전체 결합]
        {extracted_text_combined[:25000]}

        위 제공된 문서 파일들과 KCSC 국가건설기준, 업로드된 부처 지침, 법정 경비 고시 기준을 정밀 대조하여 4단계 등급화 판정 보고서를 마크다운 형식으로 작성해 주십시오.
        """

        content_payload = []
        for gfile in uploaded_genai_files:
            content_payload.append(gemini_file_part(gfile))
        content_payload.append(system_instruction + "\n\n" + prompt)
        
        ai_result = ""
        for model_name in models_to_try:
            try:
                print(f"[Design Review] Trying model: {model_name}")
                model = genai.GenerativeModel(model_name=model_name)
                response = model.generate_content(content_payload)
                ai_result = response.text
                print(f"[Design Review] Success with model: {model_name}")
                break
            except Exception as e:
                print(f"[Design Review] Failed with model {model_name}: {e}")
                continue
                
        if not ai_result:
            ai_result = "🔴 [오류]: 첨부파일 및 AI 분석 과정에서 오류가 발생했습니다. 잠시 후 다시 시도해 주세요."
            
        JOBS[job_id] = {
            "status": "completed",
            "result": ai_result
        }
    except Exception as e:
        print(f"run_design_review error: {e}")
        JOBS[job_id] = {
            "status": "error",
            "message": str(e)
        }

@app.route('/api/analyze/design_review', methods=['POST'])
def api_design_review():
    try:
        project_name = request.form.get('projectName', '설계도서 검토 프로젝트')
        project_domain = request.form.get('projectDomain', 'civil')
        review_modes_str = request.form.get('reviewModes', '[]')
        try:
            review_modes = json.loads(review_modes_str)
        except:
            review_modes = []
        additional_notes = request.form.get('additionalNotes', '')
        
        saved_files = {}
        import tempfile

        file_idx = 0
        for key in ['file_report', 'file_estimate', 'file_quantity', 'file_drawing']:
            file_objs = request.files.getlist(key)
            for file_obj in file_objs:
                if file_obj and file_obj.filename:
                    fname = file_obj.filename
                    ext = os.path.splitext(fname)[1]
                    fd, temp_path = tempfile.mkstemp(suffix=ext)
                    os.close(fd)
                    file_obj.save(temp_path)
                    saved_files[f"{key}_{file_idx}"] = {
                        'path': temp_path,
                        'name': fname,
                        'category': key
                    }
                    file_idx += 1

        job_id = str(uuid.uuid4())
        JOBS[job_id] = {"status": "processing"}
        
        thread = threading.Thread(
            target=run_design_review,
            args=(job_id, project_name, project_domain, review_modes, additional_notes, saved_files)
        )
        thread.daemon = True
        thread.start()
        
        return jsonify({"success": True, "jobId": job_id})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# ----------------------------------------------------------------------------
# 부처별 실무 지침 문서함 (요청: "각부처지침은 api가없으니 따로 업로드해야해")
# KCSC API처럼 자동으로 못 가져오는 각 부처/발주기관 실무 지침·매뉴얼을 사용자가
# 직접 업로드해서 쌓아두는 상시 보관함. 설계도서 검토 실행 시 자동으로 반영된다.
# ⚠️ 로컬 디스크(data/guidelines/)에 저장되므로, 배포 환경에 반영하려면 로컬에서
# 업로드 후 git commit/push가 필요하다(Cloudtype은 재배포 시 디스크 초기화).
# ----------------------------------------------------------------------------
@app.route('/api/guidelines', methods=['GET'])
def api_list_guidelines():
    try:
        from guideline_store import list_guidelines
        return jsonify({"success": True, "data": list_guidelines()})
    except Exception as e:
        return jsonify({"success": False, "message": str(e), "data": []})


@app.route('/api/guidelines', methods=['POST'])
def api_upload_guideline():
    try:
        from guideline_store import add_guideline

        file_obj = request.files.get('file')
        label = request.form.get('label', '').strip()
        if not file_obj or not file_obj.filename:
            return jsonify({"success": False, "message": "파일을 선택해주세요."}), 400

        import tempfile
        fname = file_obj.filename
        ext = os.path.splitext(fname)[1]
        fd, temp_path = tempfile.mkstemp(suffix=ext)
        os.close(fd)
        file_obj.save(temp_path)
        try:
            entry = add_guideline(temp_path, fname, label)
        finally:
            try:
                os.remove(temp_path)
            except Exception:
                pass

        return jsonify({"success": True, "data": entry})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/guidelines/<guideline_id>', methods=['DELETE'])
def api_delete_guideline(guideline_id):
    try:
        from guideline_store import delete_guideline
        ok = delete_guideline(guideline_id)
        if not ok:
            return jsonify({"success": False, "message": "해당 지침 문서를 찾을 수 없습니다."}), 404
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# 「건설공사 사업관리방식 검토기준 및 업무수행지침」 제127조(착공신고서 검토 및 보고) 원문.
# 이 문서 자체는 682KB짜리라 fetch_local_law_data의 키워드 매칭에 맡기면 제127조가 잘려
# 나가거나 아예 안 걸릴 위험이 있어서, 착공서류 검토 기능에서는 해당 조문을 직접 프롬프트에
# 박아넣는다 (조문 원문은 data/laws/건설공사_사업관리방식_검토기준_및_업무수행지침.md 1856~1866행 확인).
COMMENCEMENT_CHECKLIST_ARTICLE = """
제127조(착공신고서 검토 및 보고) 공사감독자는 건설공사가 착공된 경우에는 시공자로부터 다음 각 호의 서류가 포함된 착공신고서를 제출받아 적정성 여부를 검토하여 7일 이내에 발주청에 보고하여야 한다.
  1. 현장기술인 지정신고서(현장관리조직, 현장대리인, 품질관리자, 안전관리자, 보건관리자)
  2. 건설공사 공정예정표
  3. 품질관리계획서 또는 품질시험계획서(실착공 전에 제출 가능) - 총공사비 규모 및 공종에 따라 작성대상 여부가 갈림(건설기술 진흥법령)
  4. 공사도급 계약서 사본 및 산출내역서
  5. 착공 전 사진
  6. 현장기술인 경력사항 확인서 및 자격증 사본
  7. 안전관리계획서(실착공 전에 제출 가능) - 건설기술 진흥법령상 일정 규모/공종 대상 공사에만 작성 의무
  8. 유해ㆍ위험방지계획서(실착공 전에 제출 가능) - 산업안전보건법령상 일정 규모/공종 대상 공사에만 작성 의무
  9. 노무동원 및 장비투입 계획서
  10. 관급자재 수급계획서
"""

# 실무에서 실제로 쓰이는 착공계 보완요청 체크리스트(발주청 실무 서식 참고, 사용자 제공).
# ⚠️ 주의: 이건 어디까지나 "실무에서 흔히 요구되는 세부 항목 목록"을 파악하기 위한 참고자료일
# 뿐이고, 여기 적힌 금액 기준·대상 여부는 특정 발주청의 내부 서식/판단일 수 있어 전국 공통이라는
# 보장이 없다. AI는 이 목록을 "무엇을 확인해야 하는지"의 힌트로만 쓰고, 실제 대상여부·기준금액
# 판단은 반드시 현행 법령(법/시행령/시행규칙)과 법제처 검색 결과를 최우선 근거로 삼아야 한다.
# (이 목록의 수치와 현행 법령이 다르면 법령이 맞다.)
COMMENCEMENT_PRACTICAL_REFERENCE = """
[참고용 - 실무 착공계 보완요청 세부 항목 예시 (금액·대상기준은 반드시 현행 법령으로 재검증할 것)]
1. 착공신고서 (감독관 경유)
2. 현장대리인(안전관리자) 신고서
3. 현장대리인(안전관리자) 재직증명서
4. 현장대리인 기술수첩 사본, 경력증명서 (예: 토목 초급기술자 3년 이상 등 등급별 경력기준 - 건설기술 진흥법령 확인)
5. 예정공정표
6. 착공내역서(산출내역서)
7. 안전보건관리계획서 (중대재해 처벌 등에 관한 법률 - 건설기술 진흥법상 "안전관리계획서"와는 별개의 문서이므로 혼동하지 말 것)
8. 산업안전보건관리비 사용계획서 (감독관 경유)
9. 환경보전비 사용계획서 (감독관 경유)
10. 품질관리(시험)비 계획서 (참고 기준: 토목 5억원 이상, 전문 2억원 이상 - 반드시 건설기술 진흥법령 원문으로 재확인)
11. 안전관리계획서 (건설기술 진흥법 - 안전관리비 계상 대상 공사인 경우 착공 시 계획서 제출)
12. 직접시공계획서, 예정공정표, 내역서 (전문공사(전기·소방·문화재 등)는 제출대상 아니고 종합공사만 대상인 경우가 있음 - 확인 필요)
13. 인원·장비 동원계획서 (참고 기준: 일정 금액 이상 - 재검증 필요)
14. 퇴직공제가입증명서 (건설근로자 고용개선법 - 참고 기준: 공사예정금액 1억원 이상)
15. 기술지도계약서 (재해예방 전문지도기관 - 산업안전보건법령, 참고 기준: 1억원 이상)
16. 건설기계대여보증서 (건설기계관리법 - 참고 기준: 1억원 이상 또는 5개월 이상 공사, 보증서 발급대상 여부 확인 필요)
17. 고용·산재보험 가입 증명서
18. 노무비 구분관리 및 지급확인제 관련 서류 (건설산업기본법 - 대상/제외 여부에 따라 합의서 또는 적용제외 확인서)
"""


def run_commencement_review(job_id, project_name, contract_amount, total_cost, additional_notes, saved_files):
    try:
        import os
        import google.generativeai as genai
        genai.configure(api_key=GEMINI_KEY)

        extracted_text_combined = ""
        uploaded_genai_files = []

        for file_key, file_info in saved_files.items():
            path = file_info['path']
            fname = file_info['name']
            if os.path.exists(path):
                try:
                    gfile = upload_file_for_gemini(path, display_name=fname)
                    uploaded_genai_files.append(gfile)
                except Exception as e:
                    print(f"genai upload failed for {fname}: {e}")

                try:
                    if path.endswith('.xlsx') or path.endswith('.xls'):
                        import openpyxl
                        wb = openpyxl.load_workbook(path, data_only=True)
                        for sheet in wb.sheetnames:
                            ws = wb[sheet]
                            extracted_text_combined += f"\n[엑셀 시트: {sheet}]\n"
                            for row in ws.iter_rows(values_only=True):
                                row_str = " | ".join([str(c) for c in row if c is not None])
                                if row_str.strip():
                                    extracted_text_combined += row_str + "\n"
                    elif path.endswith('.pdf'):
                        import fitz
                        doc = fitz.open(path)
                        for page in doc:
                            extracted_text_combined += page.get_text() + "\n"
                        doc.close()
                    else:
                        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                            extracted_text_combined += f.read() + "\n"
                except Exception as e:
                    print(f"Text extraction failed for {fname}: {e}")

                try:
                    os.remove(path)
                except:
                    pass

        # 3개 조건부 서류(품질관리계획서/안전관리계획서/유해위험방지계획서) 대상 여부 판단을 돕기 위해
        # 법제처 실시간 검색 컨텍스트를 확보 (없으면 프롬프트 지시대로 "확인 필요"로 처리됨)
        from mcp_agent_sync import get_mcp_context_sync
        mcp_query = (
            f"공사금액(도급액) {contract_amount}원, 총공사비 {total_cost}원 규모의 건설공사에서 "
            f"안전관리계획서, 안전보건관리계획서(중대재해처벌법), 유해위험방지계획서, "
            f"품질관리계획서(또는 품질시험계획서), 퇴직공제가입, 기술지도계약, 건설기계대여보증, "
            f"노무비구분관리 대상 및 작성 의무 기준 금액"
        )
        moleg_context = get_mcp_context_sync(mcp_query)

        try:
            available_models = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods and 'vision' not in m.name.lower()]
            models_to_try = sorted(available_models, key=lambda x: (0 if '1.5-pro' in x else 1 if '2.5-pro' in x else 2 if 'pro' in x else 3 if '1.5-flash' in x else 4))
        except:
            models_to_try = ['models/gemini-1.5-pro', 'models/gemini-2.5-pro', 'models/gemini-1.5-flash']

        system_instruction = """
        당신은 대한민국 공공발주 건설공사의 공사감독관(감독자) 업무를 보좌하는 최고 수준의 전문가입니다.
        「건설공사 사업관리방식 검토기준 및 업무수행지침」 제127조에 따라, 시공자가 제출한 착공신고서
        첨부서류를 검토하여 발주청 보고(제출일로부터 7일 이내) 전에 문제를 찾아내는 것이 당신의 임무입니다.

        [자료 우선순위 - 반드시 지킬 것]
        0. 🚨 [최우선] 서류 대상 여부·기준 금액·근거 조항을 판단할 때는 항상 "현행 법령(법률·시행령·
           시행규칙)"이 최우선 근거입니다. [건설공사 사업관리방식 검토기준 및 업무수행지침 - 근거 조문],
           [법제처 실시간 검색 컨텍스트], [로컬 법령 데이터베이스]에 있는 현행 법령 내용을 최우선으로
           삼으십시오. 아래 [실무 참고 체크리스트]는 어떤 서류들을 챙겨야 하는지 "빠짐없이 훑어보기
           위한 참고 목록"일 뿐이며, 그 안의 금액 기준이나 대상 여부 서술은 특정 발주청의 내부 관행일
           수 있어 전국 공통이라는 보장이 없습니다. 참고 체크리스트의 수치가 실제 법령 검색 결과와
           다르면 반드시 법령 쪽을 따르고, 그 사실을 검토의견에 명시하십시오. 법령으로 확인이 안 되면
           참고 체크리스트의 수치를 "실무 관행상 참고치이며 발주청 재확인 필요"라고 표시하여 인용만
           하고 확정 판단에는 쓰지 마십시오.

        [검토 원칙]
        1. 제127조의 10개 대분류와 [실무 참고 체크리스트]의 세부 항목을 함께 참고하여, 아래 [출력 형식]의
           18개 세부 항목 전부를 하나씩 대조하고 첨부된 서류들 중 어느 항목에 해당하는지 판단하세요.
        2. "안전보건관리계획서"(중대재해처벌법 근거)와 "안전관리계획서"(건설기술 진흥법 근거)는 근거
           법령이 다른 별개의 문서이니 절대 혼동하거나 하나로 합쳐 다루지 마십시오.
        3. 품질관리(시험)비계획서, 안전관리계획서, 유해위험방지계획서, 인원·장비 동원계획서,
           퇴직공제가입증명서, 기술지도계약서, 건설기계대여보증서, 노무비구분관리 서류는 공사금액·
           공종·기간 규모에 따라 애초에 작성 의무가 없을 수도 있는 조건부 서류입니다. 제공된
           [공사 정보]와 현행 법령 근거로 대상 여부를 판단하되, 근거가 불충분하면 절대 추측하지
           말고 반드시 "확인 필요"로 표시하고 그 이유(어떤 법령·기준을 직접 확인해야 하는지)를 적으세요.
        4. 서류가 실제로 첨부됐다면 내용의 정합성도 검토하세요 (예: 공사도급 계약서의 금액·공사기간이
           산출내역서 합계·공정예정표와 일치하는지, 현장기술인 자격증 사본과 경력확인서의 인적사항이
           일치하는지, 서명·날인 누락 여부 등).
        5. 판단 근거(법조항, 고시명 등)를 반드시 명시하고, 근거 없이 결론만 내리지 마십시오.
        6. 실무 공문서체로, 감독관이 그대로 발주청 보고서에 옮겨쓸 수 있는 수준으로 작성하세요.
        7. ⚖️ [법령 인용은 항상 링크로 - 예외 없음] 검토의견, 종합 의견, 보완 지시사항, 법령 확인이
           필요한 항목 등 본문 어디에서든 법령명이나 조항 번호를 언급할 때는 절대 맨 텍스트로 쓰지
           말고, 반드시 아래 형식의 마크다운 링크로 작성하여 클릭 가능하게 만드십시오. (표 안의
           "검토의견" 칸에서도 동일하게 적용합니다.)
           - 법률/시행령/시행규칙(예: ~법, ~법 시행령, ~법 시행규칙):
             `[법령명 제X조](https://www.law.go.kr/법령/법령명/제X조)`
           - 고시/훈령/예규/지침/기준(예: 「건설공사 사업관리방식 검토기준 및 업무수행지침」):
             `[행정규칙명 제X조](https://www.law.go.kr/행정규칙/행정규칙명/제X조)`
           (법령명·조문 사이 띄어쓰기는 그대로 유지하고, URL을 직접 인코딩하지 마십시오.)
        8. 🚨 [사업 특성 창작 절대 금지] [법제처 실시간 검색 컨텍스트]에는 법령이 정한 대상 공사의
           "종류 목록"(예: 터널 공사, 폭발물을 사용하는 공사 등)이 예시로 함께 검색될 수 있습니다.
           이건 법령상 일반적인 분류 기준일 뿐, 이번 사업이 그 종류에 해당한다는 뜻이 절대 아닙니다.
           [공사 기본 정보]의 "추가 참고사항"이나 첨부파일에 사용자가 직접 명시하지 않은 공사 특성
           (터널·발파·굴착 깊이 등)을 마치 이 사업의 실제 특성인 것처럼 단정하거나 인용하지 마십시오.
           공종/규모 정보가 부족해서 대상 여부를 판단할 수 없다면, 정직하게 "제공된 정보만으로는
           판단할 수 없으므로 실제 공종·규모를 확인해야 한다"고만 쓰십시오.

        [출력 형식 - 반드시 마크다운 표로 작성. "검토의견" 칸에 법령을 쓸 때도 위 7번 링크 형식을 지키세요]
        ### 착공신고서 검토 결과표
        | 번호 | 서류명 | 대상여부 | 제출여부 | 검토의견 |
        |---|---|---|---|---|
        (아래 18개 세부 항목 전부를 순서대로 다루되, 서류명은 실제 문맥에 맞게 다듬어도 됩니다:
        1.착공신고서(감독관 경유) 2.현장대리인(안전관리자) 신고서 3.현장대리인(안전관리자) 재직증명서
        4.현장대리인 기술수첩 사본·경력증명서 5.예정공정표 6.착공내역서(산출내역서)
        7.안전보건관리계획서(중대재해처벌법) 8.산업안전보건관리비 사용계획서(감독관 경유)
        9.환경보전비 사용계획서(감독관 경유) 10.품질관리(시험)비계획서 11.안전관리계획서(건설기술진흥법)
        12.직접시공계획서 등 13.인원·장비 동원계획서 14.퇴직공제가입증명서 15.기술지도계약서
        16.건설기계대여보증서 17.고용·산재보험 가입증명서 18.노무비구분관리 관련 서류)

        표 다음에 아래 섹션을 이어서 작성하세요:
        ### 종합 의견
        ### 보완 지시가 필요한 사항 (시공자에게 문서로 보완 요청할 항목)
        ### 발주청 보고 시 유의사항
        ### 법령 확인이 필요한 항목 (참고 체크리스트 수치와 법령 검색결과가 다르거나, 법령으로 확정 못한 항목만 모아서 명시)
        """

        prompt = f"""
        [공사 기본 정보]
        - 사업명(공사명): {project_name}
        - 도급액(공사금액): {contract_amount or '미기재'}
        - 총공사비(추정): {total_cost or '미기재'}
        - 추가 참고사항: {additional_notes or '없음'}

        [건설공사 사업관리방식 검토기준 및 업무수행지침 - 근거 조문 (법정 최우선 근거)]
        {COMMENCEMENT_CHECKLIST_ARTICLE}

        {COMMENCEMENT_PRACTICAL_REFERENCE}

        [법제처 실시간 검색 컨텍스트 - 현행 법령 확인용. ⚠️ 다만 법령이 정한 대상 공사의 "종류 목록"
        (예: 터널 공사, 폭발물을 사용하는 공사 등)이 예시로 함께 검색될 수 있는데, 이는 일반적인 분류
        기준일 뿐 이 사업의 실제 특성이 아닙니다. "터널", "발파" 등이 언급되어도 이번 사업이 그렇다는
        뜻은 아닙니다.]
        {moleg_context}

        [첨부파일에서 추출한 텍스트 일부]
        {extracted_text_combined[:25000]}

        위 자료를 바탕으로 착공신고서 검토 결과표와 종합 의견을 작성해 주십시오.
        """

        content_payload = []
        for gfile in uploaded_genai_files:
            content_payload.append(gemini_file_part(gfile))
        content_payload.append(system_instruction + "\n\n" + prompt)

        ai_result = ""
        for model_name in models_to_try:
            try:
                print(f"[Commencement Review] Trying model: {model_name}")
                model = genai.GenerativeModel(model_name=model_name)
                response = model.generate_content(content_payload)
                ai_result = response.text
                print(f"[Commencement Review] Success with model: {model_name}")
                break
            except Exception as e:
                print(f"[Commencement Review] Failed with model {model_name}: {e}")
                continue

        if not ai_result:
            ai_result = "🔴 [오류]: 첨부파일 및 AI 분석 과정에서 오류가 발생했습니다. 잠시 후 다시 시도해 주세요."

        JOBS[job_id] = {
            "status": "completed",
            "result": ai_result
        }
    except Exception as e:
        print(f"run_commencement_review error: {e}")
        JOBS[job_id] = {
            "status": "error",
            "message": str(e)
        }

@app.route('/api/analyze/commencement_review', methods=['POST'])
def api_commencement_review():
    try:
        project_name = request.form.get('projectName', '착공서류 검토')
        contract_amount = request.form.get('contractAmount', '')
        total_cost = request.form.get('totalCost', '')
        additional_notes = request.form.get('additionalNotes', '')

        saved_files = {}
        import tempfile

        file_idx = 0
        for file_obj in request.files.getlist('files'):
            if file_obj and file_obj.filename:
                fname = file_obj.filename
                ext = os.path.splitext(fname)[1]
                fd, temp_path = tempfile.mkstemp(suffix=ext)
                os.close(fd)
                file_obj.save(temp_path)
                saved_files[f"file_{file_idx}"] = {
                    'path': temp_path,
                    'name': fname
                }
                file_idx += 1

        if not saved_files:
            return jsonify({"success": False, "message": "착공계 관련 서류를 하나 이상 첨부해주세요."}), 400

        job_id = str(uuid.uuid4())
        JOBS[job_id] = {"status": "processing"}

        thread = threading.Thread(
            target=run_commencement_review,
            args=(job_id, project_name, contract_amount, total_cost, additional_notes, saved_files)
        )
        thread.daemon = True
        thread.start()

        return jsonify({"success": True, "jobId": job_id})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)
