"""여러 문서 형식에서 텍스트를 뽑아내는 공용 유틸리티.

.xlsx/.pdf는 각각 openpyxl/PyMuPDF로 제대로 파싱하고, 그 외 확장자(.docx,
.hwp 등)는 UTF-8 텍스트로 최선을 다해 시도한다. 바이너리 형식이라 이 방식
으로는 내용이 거의 안 읽힐 수 있지만, 설계도서 검토 기능에서는 원본 파일을
Gemini Files API에도 함께 올려서 AI가 직접 읽게 하므로 완전히 못 쓰는
자료가 되지는 않는다. (원래 app.py의 run_design_review/run_commencement_review
안에 각각 따로 있던 동일한 추출 로직을 하나로 합친 것 - 지침 문서함
(guideline_store.py)에서도 그대로 재사용한다.)
"""


def extract_text_from_file(path: str) -> str:
    text = ""
    try:
        if path.endswith('.xlsx') or path.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text += f"\n[엑셀 시트: {sheet}]\n"
                for row in ws.iter_rows(values_only=True):
                    row_str = " | ".join([str(c) for c in row if c is not None])
                    if row_str.strip():
                        text += row_str + "\n"
        elif path.endswith('.pdf'):
            import fitz
            doc = fitz.open(path)
            for page in doc:
                text += page.get_text() + "\n"
            doc.close()
        else:
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                text += f.read() + "\n"
    except Exception as e:
        print(f"[doc_extract] Text extraction failed for {path}: {e}")
    return text
